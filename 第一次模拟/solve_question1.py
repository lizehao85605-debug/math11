"""Solve Question 1 of the logistics sorting scheduling problem.

Model assumptions:
1. A shift starts on an integer hour and lasts 8 consecutive hours.
2. A shift must stay within the same day, so candidate starts are 0..16.
3. Exactly five distinct shift start times are selected each day.
4. Shifts may overlap. Each worker is assigned to exactly one shift per day.
5. Freight may wait within the day, but cannot be processed before arrival and
   all freight must be processed by the end of hour 23.
6. Each worker processes at most 25 items per hour.

The program uses lexicographic optimization for each day:
1. Minimize total workers.
2. With the minimum worker count fixed, minimize peak hourly workers.
3. With both fixed, minimize hourly staffing deviation from the daily average.
4. With the first three objectives fixed, minimize total hourly backlog.
"""

from __future__ import annotations

import argparse
import math
from dataclasses import dataclass
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.optimize import Bounds, LinearConstraint, milp


HOURS = 24
SHIFT_LENGTH = 8
STARTS = tuple(range(HOURS - SHIFT_LENGTH + 1))  # 0, ..., 16
RATE = 25


@dataclass(frozen=True)
class VariableIndex:
    b: np.ndarray
    x: np.ndarray
    y: np.ndarray
    inventory: np.ndarray
    deviation: np.ndarray
    peak: int
    size: int


@dataclass
class DaySolution:
    day: int
    arrivals: np.ndarray
    selected: np.ndarray
    workers_by_start: np.ndarray
    processed: np.ndarray
    inventory: np.ndarray
    hourly_workers: np.ndarray
    total_workers: int
    peak_workers: int
    balance_deviation: float
    theoretical_lower_bound: int


def make_indices() -> VariableIndex:
    offset = 0
    b = np.arange(offset, offset + len(STARTS))
    offset += len(STARTS)
    x = np.arange(offset, offset + len(STARTS))
    offset += len(STARTS)
    y = np.arange(offset, offset + HOURS)
    offset += HOURS
    inventory = np.arange(offset, offset + HOURS)
    offset += HOURS
    deviation = np.arange(offset, offset + HOURS)
    offset += HOURS
    peak = offset
    offset += 1
    return VariableIndex(
        b=b,
        x=x,
        y=y,
        inventory=inventory,
        deviation=deviation,
        peak=peak,
        size=offset,
    )


IDX = make_indices()


def shift_coverage() -> np.ndarray:
    """Return A[h, t] = 1 if the shift starting at t covers hour h."""
    return np.array(
        [[int(start <= hour < start + SHIFT_LENGTH) for start in STARTS] for hour in range(HOURS)],
        dtype=float,
    )


COVERAGE = shift_coverage()


def load_arrivals(path: Path) -> dict[int, np.ndarray]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.active
    raw: dict[int, dict[int, float]] = {}

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(value is None for value in row[:3]):
            continue
        day, hour, quantity = row[:3]
        if day is None or hour is None or quantity is None:
            raise ValueError(f"Row {row_number} has an empty day, hour, or quantity.")
        day = int(day)
        hour = int(hour)
        quantity = float(quantity)
        if not 0 <= hour < HOURS:
            raise ValueError(f"Row {row_number} has invalid hour {hour}.")
        if quantity < 0:
            raise ValueError(f"Row {row_number} has negative quantity {quantity}.")
        if hour in raw.setdefault(day, {}):
            raise ValueError(f"Duplicate record for day {day}, hour {hour}.")
        raw[day][hour] = quantity

    if not raw:
        raise ValueError("No arrival records were found in the workbook.")

    arrivals: dict[int, np.ndarray] = {}
    for day in sorted(raw):
        missing = sorted(set(range(HOURS)) - set(raw[day]))
        if missing:
            raise ValueError(f"Day {day} is missing hours: {missing}")
        arrivals[day] = np.array([raw[day][hour] for hour in range(HOURS)], dtype=float)
    return arrivals


def append_constraint(
    rows: list[np.ndarray],
    lower: list[float],
    upper: list[float],
    coefficients: dict[int, float],
    lb: float,
    ub: float,
) -> None:
    row = np.zeros(IDX.size, dtype=float)
    for index, value in coefficients.items():
        row[index] = value
    rows.append(row)
    lower.append(lb)
    upper.append(ub)


def build_base_model(arrivals: np.ndarray) -> tuple[list[np.ndarray], list[float], list[float], Bounds, np.ndarray]:
    rows: list[np.ndarray] = []
    lower: list[float] = []
    upper: list[float] = []

    # Exactly five distinct shift start times are selected.
    append_constraint(rows, lower, upper, {int(i): 1.0 for i in IDX.b}, 5.0, 5.0)

    # Big-M linking: b[t] <= x[t] <= M*b[t].
    # The chosen M can process the entire daily volume in a single hour, so it
    # cannot restrict a feasible staffing solution.
    big_m = max(1, math.ceil(float(arrivals.sum()) / RATE))
    for position in range(len(STARTS)):
        b_index = int(IDX.b[position])
        x_index = int(IDX.x[position])
        append_constraint(rows, lower, upper, {x_index: 1.0, b_index: -float(big_m)}, -np.inf, 0.0)
        append_constraint(rows, lower, upper, {b_index: 1.0, x_index: -1.0}, -np.inf, 0.0)

    # Hourly processing cannot exceed the active workers' capacity.
    for hour in range(HOURS):
        coefficients = {int(IDX.y[hour]): 1.0}
        for position in range(len(STARTS)):
            if COVERAGE[hour, position]:
                coefficients[int(IDX.x[position])] = -float(RATE)
        append_constraint(rows, lower, upper, coefficients, -np.inf, 0.0)

    # Inventory balance. Nonnegative inventory prevents processing freight
    # before it has arrived.
    for hour in range(HOURS):
        coefficients = {
            int(IDX.inventory[hour]): 1.0,
            int(IDX.y[hour]): 1.0,
        }
        if hour > 0:
            coefficients[int(IDX.inventory[hour - 1])] = -1.0
        append_constraint(
            rows,
            lower,
            upper,
            coefficients,
            float(arrivals[hour]),
            float(arrivals[hour]),
        )

    # All freight must be processed by the end of the day.
    append_constraint(
        rows,
        lower,
        upper,
        {int(IDX.inventory[HOURS - 1]): 1.0},
        0.0,
        0.0,
    )

    # Peak workers is at least the number of active workers in every hour.
    for hour in range(HOURS):
        coefficients = {IDX.peak: -1.0}
        for position in range(len(STARTS)):
            if COVERAGE[hour, position]:
                coefficients[int(IDX.x[position])] = 1.0
        append_constraint(rows, lower, upper, coefficients, -np.inf, 0.0)

    variable_lower = np.zeros(IDX.size, dtype=float)
    variable_upper = np.full(IDX.size, np.inf, dtype=float)
    variable_upper[IDX.b] = 1.0
    variable_upper[IDX.x] = float(big_m)
    bounds = Bounds(variable_lower, variable_upper)

    integrality = np.zeros(IDX.size, dtype=int)
    integrality[IDX.b] = 1
    integrality[IDX.x] = 1
    integrality[IDX.deviation] = 1
    integrality[IDX.peak] = 1
    return rows, lower, upper, bounds, integrality


def run_milp(
    objective: np.ndarray,
    rows: list[np.ndarray],
    lower: list[float],
    upper: list[float],
    bounds: Bounds,
    integrality: np.ndarray,
    time_limit: float,
    stage: str,
    day: int,
):
    matrix = np.vstack(rows)
    constraints = LinearConstraint(matrix, np.array(lower), np.array(upper))
    result = milp(
        c=objective,
        integrality=integrality,
        bounds=bounds,
        constraints=constraints,
        options={
            "disp": False,
            "presolve": True,
            "time_limit": time_limit,
            "mip_rel_gap": 0.0,
        },
    )
    if result.status != 0 or result.x is None:
        raise RuntimeError(
            f"Day {day}, {stage} did not reach an optimum. "
            f"status={result.status}, message={result.message}"
        )
    return result


def solve_day(day: int, arrivals: np.ndarray, time_limit: float) -> DaySolution:
    rows, lower, upper, bounds, integrality = build_base_model(arrivals)

    # Stage 1: minimize total distinct workers used that day.
    objective = np.zeros(IDX.size)
    objective[IDX.x] = 1.0
    stage1 = run_milp(objective, rows, lower, upper, bounds, integrality, time_limit, "stage 1", day)
    optimum_workers = int(round(float(stage1.x[IDX.x].sum())))

    worker_row = {int(index): 1.0 for index in IDX.x}
    append_constraint(rows, lower, upper, worker_row, optimum_workers, optimum_workers)

    # Stage 2: among minimum-worker solutions, minimize peak hourly staffing.
    objective = np.zeros(IDX.size)
    objective[IDX.peak] = 1.0
    stage2 = run_milp(objective, rows, lower, upper, bounds, integrality, time_limit, "stage 2", day)
    optimum_peak = int(round(float(stage2.x[IDX.peak])))

    append_constraint(rows, lower, upper, {IDX.peak: 1.0}, optimum_peak, optimum_peak)

    # Stage 3: among the two previous optima, balance hourly staffing. Instead
    # of using the fractional average N/3 directly, deviation[h] models
    # |3*n[h] - N|. Minimizing its sum is exactly equivalent to minimizing the
    # sum of absolute deviations from the average N/3.
    for hour in range(HOURS):
        above_average = {int(IDX.deviation[hour]): 1.0}
        below_average = {int(IDX.deviation[hour]): 1.0}
        for position in range(len(STARTS)):
            if COVERAGE[hour, position]:
                above_average[int(IDX.x[position])] = -3.0
                below_average[int(IDX.x[position])] = 3.0
        append_constraint(
            rows,
            lower,
            upper,
            above_average,
            -float(optimum_workers),
            np.inf,
        )
        append_constraint(
            rows,
            lower,
            upper,
            below_average,
            float(optimum_workers),
            np.inf,
        )

    objective = np.zeros(IDX.size)
    objective[IDX.deviation] = 1.0
    stage3 = run_milp(objective, rows, lower, upper, bounds, integrality, time_limit, "stage 3", day)
    optimum_deviation_scaled = int(round(float(stage3.x[IDX.deviation].sum())))

    deviation_row = {int(index): 1.0 for index in IDX.deviation}
    append_constraint(
        rows,
        lower,
        upper,
        deviation_row,
        optimum_deviation_scaled,
        optimum_deviation_scaled,
    )

    # Stage 4: with worker count, peak staffing, and staffing balance fixed,
    # process freight as early as possible.
    objective = np.zeros(IDX.size)
    objective[IDX.inventory] = 1.0
    stage4 = run_milp(objective, rows, lower, upper, bounds, integrality, time_limit, "stage 4", day)

    values = stage4.x
    selected = np.rint(values[IDX.b]).astype(int)
    workers = np.rint(values[IDX.x]).astype(int)
    processed = values[IDX.y].copy()
    inventory = values[IDX.inventory].copy()
    hourly_workers = np.rint(COVERAGE @ workers).astype(int)

    solution = DaySolution(
        day=day,
        arrivals=arrivals,
        selected=selected,
        workers_by_start=workers,
        processed=processed,
        inventory=inventory,
        hourly_workers=hourly_workers,
        total_workers=int(workers.sum()),
        peak_workers=int(hourly_workers.max()),
        balance_deviation=float(np.abs(3 * hourly_workers - optimum_workers).sum()) / 3.0,
        theoretical_lower_bound=math.ceil(float(arrivals.sum()) / (SHIFT_LENGTH * RATE)),
    )
    validate_solution(solution)
    return solution


def validate_solution(solution: DaySolution, tolerance: float = 1e-5) -> None:
    if int(solution.selected.sum()) != 5:
        raise AssertionError(f"Day {solution.day}: selected {solution.selected.sum()} shifts, expected 5.")
    if np.any((solution.selected == 0) & (solution.workers_by_start != 0)):
        raise AssertionError(f"Day {solution.day}: workers assigned to an unselected shift.")
    if np.any((solution.selected == 1) & (solution.workers_by_start < 1)):
        raise AssertionError(f"Day {solution.day}: a selected shift has no workers.")
    if np.any(solution.processed < -tolerance) or np.any(solution.inventory < -tolerance):
        raise AssertionError(f"Day {solution.day}: negative processing or inventory.")
    if np.any(solution.processed - RATE * solution.hourly_workers > tolerance):
        raise AssertionError(f"Day {solution.day}: hourly capacity exceeded.")

    reconstructed = np.empty(HOURS)
    previous = 0.0
    for hour in range(HOURS):
        reconstructed[hour] = previous + solution.arrivals[hour] - solution.processed[hour]
        previous = reconstructed[hour]
    if not np.allclose(reconstructed, solution.inventory, atol=tolerance):
        raise AssertionError(f"Day {solution.day}: inventory balance failed.")
    if abs(solution.inventory[-1]) > tolerance:
        raise AssertionError(f"Day {solution.day}: end-of-day inventory is not zero.")
    if solution.total_workers != int(solution.workers_by_start.sum()):
        raise AssertionError(f"Day {solution.day}: inconsistent total worker count.")
    if solution.peak_workers != int(solution.hourly_workers.max()):
        raise AssertionError(f"Day {solution.day}: inconsistent peak worker count.")


def display_number(value: float):
    rounded = round(float(value))
    if abs(float(value) - rounded) < 1e-6:
        return int(rounded)
    return round(float(value), 4)


def style_worksheet(worksheet, widths: dict[int, float]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = width


def write_results(path: Path, solutions: list[DaySolution]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "每日汇总"
    summary.append(
        [
            "天",
            "全天货量",
            "200件理论下界",
            "最优总人数",
            "峰值在岗人数",
            "平均在岗人数",
            "小时人数总绝对偏差",
            "下界差值",
            "选中班次",
        ]
    )

    shifts = workbook.create_sheet("班次安排")
    shifts.append(["天", "班次序号", "开始时间", "结束时间", "工人数"])

    hourly = workbook.create_sheet("逐小时详情")
    hourly.append(
        ["天", "小时", "进货量", "处理量", "剩余货量", "在岗人数", "相对平均人数偏差", "最大处理能力"]
    )

    matrix = workbook.create_sheet("班次人数矩阵")
    matrix.append(["天"] + [f"{start}:00-{start + SHIFT_LENGTH}:00" for start in STARTS] + ["总人数"])

    for solution in solutions:
        selected_positions = np.flatnonzero(solution.selected)
        shift_labels = [f"{STARTS[pos]:02d}:00-{STARTS[pos] + SHIFT_LENGTH:02d}:00" for pos in selected_positions]
        summary.append(
            [
                solution.day,
                display_number(solution.arrivals.sum()),
                solution.theoretical_lower_bound,
                solution.total_workers,
                solution.peak_workers,
                round(solution.total_workers / 3.0, 4),
                round(solution.balance_deviation, 4),
                solution.total_workers - solution.theoretical_lower_bound,
                "；".join(shift_labels),
            ]
        )

        for order, position in enumerate(selected_positions, start=1):
            start = STARTS[position]
            shifts.append([solution.day, order, start, start + SHIFT_LENGTH, int(solution.workers_by_start[position])])

        for hour in range(HOURS):
            hourly.append(
                [
                    solution.day,
                    hour,
                    display_number(solution.arrivals[hour]),
                    display_number(solution.processed[hour]),
                    display_number(solution.inventory[hour]),
                    int(solution.hourly_workers[hour]),
                    round(abs(solution.hourly_workers[hour] - solution.total_workers / 3.0), 4),
                    int(RATE * solution.hourly_workers[hour]),
                ]
            )

        matrix.append([solution.day] + [int(value) for value in solution.workers_by_start] + [solution.total_workers])

    style_worksheet(summary, {1: 8, 2: 14, 3: 16, 4: 14, 5: 16, 6: 14, 7: 20, 8: 12, 9: 55})
    style_worksheet(shifts, {1: 8, 2: 12, 3: 12, 4: 12, 5: 12})
    style_worksheet(hourly, {1: 8, 2: 10, 3: 14, 4: 14, 5: 14, 6: 12, 7: 20, 8: 16})
    style_worksheet(matrix, {1: 8, **{i: 15 for i in range(2, 19)}, 19: 12})

    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Solve Question 1 using a mixed-integer linear program.")
    parser.add_argument("--input", type=Path, default=Path("第一轮B题附件.xlsx"), help="Input Excel file")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("问题一优化结果_人数均衡.xlsx"),
        help="Output Excel file",
    )
    parser.add_argument(
        "--days",
        type=int,
        nargs="*",
        help="Optional day numbers to solve; by default all days are solved",
    )
    parser.add_argument(
        "--time-limit",
        type=float,
        default=60.0,
        help="MILP time limit in seconds for each optimization stage",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    arrivals_by_day = load_arrivals(args.input)
    selected_days = sorted(args.days) if args.days else sorted(arrivals_by_day)
    unknown_days = [day for day in selected_days if day not in arrivals_by_day]
    if unknown_days:
        raise ValueError(f"The input file does not contain days: {unknown_days}")

    solutions: list[DaySolution] = []
    for day in selected_days:
        solution = solve_day(day, arrivals_by_day[day], args.time_limit)
        solutions.append(solution)
        selected_starts = [STARTS[pos] for pos in np.flatnonzero(solution.selected)]
        print(
            f"day={day:02d} workers={solution.total_workers} "
            f"peak={solution.peak_workers} balance={solution.balance_deviation:.2f} "
            f"starts={selected_starts}"
        )

    write_results(args.output, solutions)
    print(f"Saved {len(solutions)} day(s) to: {args.output.resolve()}")


if __name__ == "__main__":
    main()
