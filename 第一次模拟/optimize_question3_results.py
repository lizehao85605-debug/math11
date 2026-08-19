"""Generate enhanced Question 3 schedules without overwriting prior results.

Two variants are produced:
1. Efficiency first: minimum hires, then minimum maximum daily surplus.
2. Fairness option: cap every work streak at six days, then minimize surplus.

Both variants re-optimize all daily shifts under Question 2 conditions and
decompose the aggregate 10-item-hour counts into named-worker assignments.
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import Bounds

from solve_question2 import STARTS, load_arrivals
from solve_question3 import (
    MAX_CONSECUTIVE,
    WORK_DAYS_PER_WORKER,
    HiringSolution,
    assign_workers_to_joint_shifts,
    build_hiring_model,
    decompose_worker_days,
    derive_daily_minimums,
    lower_bounds,
    maximum_consecutive,
    solve_hiring,
    solve_milp,
    solve_question3_daily_schedules,
    style_worksheet,
    validate_aggregate_solution,
    write_joint_results,
)


def solve_hiring_with_streak_limit(
    daily_minimum: np.ndarray,
    time_limit: float,
    streak_limit: int,
) -> HiringSolution:
    """Apply a stricter streak cap while preserving the original flow model."""
    if streak_limit == MAX_CONSECUTIVE:
        return solve_hiring(daily_minimum, time_limit)
    if not 1 <= streak_limit < MAX_CONSECUTIVE:
        raise ValueError(f"streak_limit must be between 1 and {MAX_CONSECUTIVE}.")

    idx, constraints, original_bounds, integrality = build_hiring_model(daily_minimum)
    lower = np.array(original_bounds.lb, copy=True)
    upper = np.array(original_bounds.ub, copy=True)

    # A transition out of state c creates a work streak of c + 1 days.
    for consecutive in range(streak_limit, MAX_CONSECUTIVE):
        upper[idx.work[:, consecutive, :].ravel()] = 0.0
    bounds = Bounds(lower, upper)

    objective = np.zeros(idx.size)
    objective[idx.hired] = 1.0
    stage1 = solve_milp(
        objective,
        constraints,
        bounds,
        integrality,
        time_limit,
        f"Hiring stage with streak cap {streak_limit}",
    )
    optimum_hired = int(round(float(stage1.x[idx.hired])))
    constraints.add({idx.hired: 1.0}, optimum_hired, optimum_hired)

    objective = np.zeros(idx.size)
    objective[idx.max_surplus] = 1.0
    stage2 = solve_milp(
        objective,
        constraints,
        bounds,
        integrality,
        time_limit,
        f"Surplus stage with streak cap {streak_limit}",
    )
    optimum_surplus = int(round(float(stage2.x[idx.max_surplus])))
    constraints.add({idx.max_surplus: 1.0}, optimum_surplus, optimum_surplus)

    objective = np.zeros(idx.size)
    for day in range(1, len(daily_minimum) + 1):
        objective[idx.population[day, streak_limit, :]] = 1.0
    stage3 = solve_milp(
        objective,
        constraints,
        bounds,
        integrality,
        time_limit,
        f"Streak tie-breaker with cap {streak_limit}",
    )

    values = stage3.x
    solution = HiringSolution(
        hired=optimum_hired,
        max_surplus=optimum_surplus,
        daily_work_counts=np.rint(values[idx.work]).astype(int).sum(axis=(1, 2)),
        population=np.rint(values[idx.population]).astype(int),
        work_flows=np.rint(values[idx.work]).astype(int),
        rest_flows=np.rint(values[idx.rest]).astype(int),
    )
    validate_aggregate_solution(solution, daily_minimum)
    return solution


def assign_individual_low_hours(
    calendar: dict[int, list[str]],
    days: list[int],
    daily_solutions: dict[int, object],
) -> dict[tuple[int, int], tuple[int, int]]:
    """Decompose shift-level low-hour counts into fair named assignments."""
    assignments: dict[tuple[int, int], tuple[int, int]] = {}
    global_history: dict[int, Counter[int]] = defaultdict(Counter)
    local_history: dict[int, Counter[int]] = defaultdict(Counter)

    for day_index, day in enumerate(days):
        solution = daily_solutions[day]
        for position in np.flatnonzero(solution.selected):
            start = STARTS[position]
            end = start + 8
            shift_label = f"{start:02d}-{end:02d}"
            members = sorted(
                worker for worker, entries in calendar.items() if entries[day_index] == shift_label
            )
            expected = int(solution.workers_by_start[position])
            if len(members) != expected:
                raise AssertionError(
                    f"Day {day}, shift {shift_label}: expected {expected} members, found {len(members)}."
                )

            remaining = set(members)
            buckets = [
                (local_hour, int(count))
                for local_hour, count in enumerate(solution.low_hour_workers[position])
                if int(count) > 0
            ]
            for local_hour, count in sorted(buckets, key=lambda item: (-item[1], item[0])):
                global_hour = start + local_hour
                candidates = sorted(
                    remaining,
                    key=lambda worker: (
                        global_history[worker][global_hour],
                        local_history[worker][local_hour],
                        worker,
                    ),
                )
                chosen = candidates[:count]
                if len(chosen) != count:
                    raise AssertionError(f"Day {day}, shift {shift_label}: low-hour decomposition failed.")
                for worker in chosen:
                    assignments[(worker, day)] = (global_hour, local_hour + 1)
                    global_history[worker][global_hour] += 1
                    local_history[worker][local_hour] += 1
                    remaining.remove(worker)

            if remaining:
                raise AssertionError(f"Day {day}, shift {shift_label}: workers lack a 10-item hour.")

    expected_assignments = sum(
        sum(entry != "\u4f11" for entry in entries) for entries in calendar.values()
    )
    if len(assignments) != expected_assignments:
        raise AssertionError("Individual low-hour assignment count is inconsistent.")
    return assignments


def enhance_workbook(
    path: Path,
    variant_name: str,
    streak_limit: int,
    daily_minimum: np.ndarray,
    minimum_solutions: dict[int, object],
    hiring: HiringSolution,
    daily_solutions: dict[int, object],
    calendar: dict[int, list[str]],
    low_assignments: dict[tuple[int, int], tuple[int, int]],
) -> None:
    workbook = load_workbook(path)
    summary = workbook.worksheets[0]
    worker_stats = workbook.worksheets[5]
    details = workbook.worksheets[7]

    actual_max_streak = max(
        maximum_consecutive([entry != "\u4f11" for entry in entries])
        for entries in calendar.values()
    )
    peak = max(solution.peak_workers for solution in daily_solutions.values())
    peak_sum = sum(solution.peak_workers for solution in daily_solutions.values())
    daily_peak_reference = sum(solution.peak_workers for solution in minimum_solutions.values())
    previous_surplus_cap = hiring.max_surplus - 1
    previous_cap_worker_days = sum(
        min(hiring.hired, int(minimum) + previous_surplus_cap)
        for minimum in daily_minimum
    )
    required_worker_days = hiring.hired * WORK_DAYS_PER_WORKER
    if previous_cap_worker_days < required_worker_days:
        surplus_certificate = (
            f"{previous_cap_worker_days}<{required_worker_days}, "
            "so the current surplus bound is necessary"
        )
    else:
        surplus_certificate = (
            f"Capacity alone is inconclusive; the full {streak_limit}-day-cap MILP proves optimality"
        )
    late_shift_counts = [
        sum(entry.startswith("16-") for entry in entries)
        for entries in calendar.values()
    ]

    enhanced_rows = [
        ("\u65b9\u6848\u7c7b\u578b", variant_name, "\u4e0d\u8986\u76d6\u539f\u7ed3\u679c"),
        ("\u8fde\u7eed\u5de5\u4f5c\u4e0a\u9650", streak_limit, "\u6a21\u578b\u4e2d\u5b9e\u9645\u65bd\u52a0\u7684\u4e0a\u9650"),
        ("\u5b9e\u9645\u6700\u5927\u8fde\u7eed\u5de5\u4f5c\u5929\u6570", actual_max_streak, "\u9010\u4eba\u6708\u5386\u8ba1\u7b97"),
        ("\u5168\u6708\u6700\u5927\u5c0f\u65f6\u5728\u5c97\u4eba\u6570", peak, "\u7b2c13\u5929\u72ec\u7acb\u4e0b\u754c\u4e3a295"),
        ("30\u5929\u6bcf\u65e5\u5cf0\u503c\u4e4b\u548c", peak_sum, "\u9010\u65e5\u91cd\u65b0\u4f18\u5316"),
        ("\u9010\u65e5\u5cf0\u503c\u4e0b\u754c\u4e4b\u548c", daily_peak_reference, "\u5df2\u9010\u65e5\u72ec\u7acb\u9a8c\u8bc1"),
        ("\u524d\u4e00\u51a7\u4f59\u4e0a\u9650\u68c0\u9a8c", previous_cap_worker_days, surplus_certificate),
        ("16\u70b9\u73ed\u6b21\u6bcf\u4eba\u6b21\u6570\u8303\u56f4", f"{min(late_shift_counts)}-{max(late_shift_counts)}", "\u7528\u4e8e\u68c0\u67e5\u591c\u73ed\u516c\u5e73\u6027"),
        ("\u9010\u4eba10\u4ef6\u5c0f\u65f6\u6570", len(low_assignments), "\u4e0e\u603b\u4e0a\u73ed\u4eba\u65e5\u4e00\u81f4"),
    ]
    for row in enhanced_rows:
        summary.append(row)

    details.cell(1, 4, "10\u4ef6\u5c0f\u65f6")
    details.cell(1, 5, "\u73ed\u6b21\u5185\u5c0f\u65f6\u5e8f\u53f7")
    worker_low_counts: Counter[int] = Counter()
    for row in range(2, details.max_row + 1):
        worker = int(details.cell(row, 1).value)
        day = int(details.cell(row, 2).value)
        global_hour, local_hour = low_assignments[(worker, day)]
        details.cell(row, 4, global_hour)
        details.cell(row, 5, local_hour)
        worker_low_counts[worker] += 1

    worker_stats.cell(1, 6, "10\u4ef6\u5c0f\u65f6\u6570")
    for row in range(2, worker_stats.max_row + 1):
        worker = int(worker_stats.cell(row, 1).value)
        worker_stats.cell(row, 6, worker_low_counts[worker])
        if worker_low_counts[worker] != WORK_DAYS_PER_WORKER:
            raise AssertionError(f"Worker {worker}: low-hour count is not 23.")

    style_worksheet(summary, {1: 31, 2: 18, 3: 46})
    style_worksheet(worker_stats, {1: 12, 2: 12, 3: 12, 4: 18, 5: 12, 6: 16})
    style_worksheet(details, {1: 12, 2: 8, 3: 12, 4: 14, 5: 18})
    workbook.save(path)


def generate_variant(
    output_path: Path,
    variant_name: str,
    streak_limit: int,
    arrivals_by_day: dict[int, np.ndarray],
    demand,
    minimum_solutions: dict[int, object],
    time_limit: float,
) -> dict[str, int]:
    hiring = solve_hiring_with_streak_limit(demand.daily_minimum, time_limit, streak_limit)
    daily_workers, work_calendar = decompose_worker_days(hiring)
    actual_max_streak = max(maximum_consecutive(flags) for flags in work_calendar[1:])
    if actual_max_streak > streak_limit:
        raise AssertionError(f"Generated roster exceeds the requested {streak_limit}-day streak cap.")

    daily_solutions = solve_question3_daily_schedules(
        arrivals_by_day,
        demand.days,
        hiring.daily_work_counts,
        time_limit,
    )
    calendar = assign_workers_to_joint_shifts(daily_workers, demand.days, daily_solutions)
    low_assignments = assign_individual_low_hours(calendar, demand.days, daily_solutions)
    write_joint_results(
        output_path,
        arrivals_by_day,
        demand,
        hiring,
        lower_bounds(demand.daily_minimum),
        daily_solutions,
        calendar,
    )
    enhance_workbook(
        output_path,
        variant_name,
        streak_limit,
        demand.daily_minimum,
        minimum_solutions,
        hiring,
        daily_solutions,
        calendar,
        low_assignments,
    )

    metrics = {
        "hired": hiring.hired,
        "max_surplus": hiring.max_surplus,
        "max_streak": actual_max_streak,
        "max_peak": max(solution.peak_workers for solution in daily_solutions.values()),
        "peak_sum": sum(solution.peak_workers for solution in daily_solutions.values()),
    }
    print(f"Saved {variant_name}: {output_path.resolve()} metrics={metrics}")
    return metrics


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate enhanced Question 3 result variants.")
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("\u7b2c\u4e00\u8f6eB\u9898\u9644\u4ef6.xlsx"),
    )
    parser.add_argument(
        "--efficiency-output",
        type=Path,
        default=Path("\u95ee\u9898\u4e09\u4f18\u5316\u7ed3\u679c_\u6548\u7387\u4f18\u5148.xlsx"),
    )
    parser.add_argument(
        "--fairness-output",
        type=Path,
        default=Path("\u95ee\u9898\u4e09\u4f18\u5316\u7ed3\u679c_\u8fde\u7eed6\u5929.xlsx"),
    )
    parser.add_argument("--time-limit", type=float, default=180.0)
    parser.add_argument("--skip-fairness", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    arrivals_by_day = load_arrivals(args.input)
    demand, minimum_solutions = derive_daily_minimums(arrivals_by_day, args.time_limit)

    efficiency = generate_variant(
        args.efficiency_output,
        "\u6548\u7387\u4f18\u5148",
        MAX_CONSECUTIVE,
        arrivals_by_day,
        demand,
        minimum_solutions,
        args.time_limit,
    )
    if efficiency != {
        "hired": 581,
        "max_surplus": 53,
        "max_streak": 7,
        "max_peak": 295,
        "peak_sum": 6498,
    }:
        raise AssertionError(f"Unexpected efficiency metrics: {efficiency}")

    if not args.skip_fairness:
        fairness = generate_variant(
            args.fairness_output,
            "\u8fde\u7eed6\u5929\u516c\u5e73\u6027",
            6,
            arrivals_by_day,
            demand,
            minimum_solutions,
            args.time_limit,
        )
        if fairness != {
            "hired": 581,
            "max_surplus": 59,
            "max_streak": 6,
            "max_peak": 295,
            "peak_sum": 6498,
        }:
            raise AssertionError(f"Unexpected fairness metrics: {fairness}")


if __name__ == "__main__":
    main()
