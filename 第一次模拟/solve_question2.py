"""Solve Question 2 of the logistics sorting scheduling problem.

Question 2 extends Question 1 with two rules:
1. Freight arriving from 00:00 through 11:59 must be completed before 16:00.
2. In every 8-hour shift, each worker processes 10 items in exactly one hour
   and 25 items in each of the other seven hours.

Model assumptions:
1. A shift starts on an integer hour, lasts 8 hours, and stays within the day.
2. Exactly five distinct starts are selected from 0..16; shifts may overlap.
3. The 10-item hour may be assigned independently for each worker within the
   worker's shift. The output reports this assignment explicitly.
4. Freight may wait after arrival. Early freight has a 16:00 deadline and all
   other freight has a 24:00 deadline.

Lexicographic objectives for each day:
1. Minimize total workers.
2. Minimize peak hourly workers with the minimum worker count fixed.
3. Minimize hourly staffing deviation from the daily average.
4. Minimize total backlog with the first three objectives fixed.
"""

from __future__ import annotations

import argparse
import math
from dataclasses import dataclass
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.optimize import Bounds, LinearConstraint, milp


HOURS = 24
EARLY_END = 12
EARLY_DEADLINE = 16
SHIFT_LENGTH = 8
STARTS = tuple(range(HOURS - SHIFT_LENGTH + 1))
NORMAL_RATE = 25
LOW_RATE = 10
RATE_LOSS = NORMAL_RATE - LOW_RATE
SHIFT_CAPACITY = (SHIFT_LENGTH - 1) * NORMAL_RATE + LOW_RATE


@dataclass(frozen=True)
class VariableIndex:
    b: np.ndarray
    x: np.ndarray
    low_hour: np.ndarray
    processed_early: np.ndarray
    processed_late: np.ndarray
    inventory_early: np.ndarray
    inventory_late: np.ndarray
    deviation: np.ndarray
    peak: int
    size: int


@dataclass
class DaySolution:
    day: int
    arrivals: np.ndarray
    selected: np.ndarray
    workers_by_start: np.ndarray
    low_hour_workers: np.ndarray
    processed_early: np.ndarray
    processed_late: np.ndarray
    inventory_early: np.ndarray
    inventory_late: np.ndarray
    hourly_workers: np.ndarray
    hourly_low_workers: np.ndarray
    hourly_capacity: np.ndarray
    total_workers: int
    peak_workers: int
    balance_deviation: float
    theoretical_lower_bound: int


def make_indices() -> VariableIndex:
    offset = 0
    b = np.arange(offset, offset + len(STARTS))
    offset += len(STARTS)
    x = np.arange(offset, offset + len(STARTS))
    offset += len(STARTS)
    low_hour = np.arange(offset, offset + len(STARTS) * SHIFT_LENGTH).reshape(len(STARTS), SHIFT_LENGTH)
    offset += len(STARTS) * SHIFT_LENGTH
    processed_early = np.arange(offset, offset + HOURS)
    offset += HOURS
    processed_late = np.arange(offset, offset + HOURS)
    offset += HOURS
    inventory_early = np.arange(offset, offset + HOURS)
    offset += HOURS
    inventory_late = np.arange(offset, offset + HOURS)
    offset += HOURS
    deviation = np.arange(offset, offset + HOURS)
    offset += HOURS
    peak = offset
    offset += 1
    return VariableIndex(
        b=b,
        x=x,
        low_hour=low_hour,
        processed_early=processed_early,
        processed_late=processed_late,
        inventory_early=inventory_early,
        inventory_late=inventory_late,
        deviation=deviation,
        peak=peak,
        size=offset,
    )


IDX = make_indices()


def shift_coverage() -> np.ndarray:
    return np.array(
        [[int(start <= hour < start + SHIFT_LENGTH) for start in STARTS] for hour in range(HOURS)],
        dtype=float,
    )


COVERAGE = shift_coverage()


def load_arrivals(path: Path) -> dict[int, np.ndarray]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.active
    raw: dict[int, dict[int, float]] = {}

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(value is None for value in row[:3]):
            continue
        day, hour, quantity = row[:3]
        if day is None or hour is None or quantity is None:
            raise ValueError(f"Row {row_number} has an empty day, hour, or quantity.")
        day = int(day)
        hour = int(hour)
        quantity = float(quantity)
        if not 0 <= hour < HOURS:
            raise ValueError(f"Row {row_number} has invalid hour {hour}.")
        if quantity < 0:
            raise ValueError(f"Row {row_number} has negative quantity {quantity}.")
        if hour in raw.setdefault(day, {}):
            raise ValueError(f"Duplicate record for day {day}, hour {hour}.")
        raw[day][hour] = quantity

    if not raw:
        raise ValueError("No arrival records were found in the workbook.")

    arrivals: dict[int, np.ndarray] = {}
    for day in sorted(raw):
        missing = sorted(set(range(HOURS)) - set(raw[day]))
        if missing:
            raise ValueError(f"Day {day} is missing hours: {missing}")
        arrivals[day] = np.array([raw[day][hour] for hour in range(HOURS)], dtype=float)
    return arrivals


def append_constraint(
    rows: list[np.ndarray],
    lower: list[float],
    upper: list[float],
    coefficients: dict[int, float],
    lb: float,
    ub: float,
) -> None:
    row = np.zeros(IDX.size, dtype=float)
    for index, value in coefficients.items():
        row[index] = value
    rows.append(row)
    lower.append(lb)
    upper.append(ub)


def build_base_model(arrivals: np.ndarray) -> tuple[list[np.ndarray], list[float], list[float], Bounds, np.ndarray]:
    rows: list[np.ndarray] = []
    lower: list[float] = []
    upper: list[float] = []

    append_constraint(rows, lower, upper, {int(i): 1.0 for i in IDX.b}, 5.0, 5.0)

    # This upper bound can process the entire daily volume in one normal-rate
    # hour and therefore cannot exclude a feasible staffing plan.
    big_m = max(1, math.ceil(float(arrivals.sum()) / NORMAL_RATE))

    for position in range(len(STARTS)):
        b_index = int(IDX.b[position])
        x_index = int(IDX.x[position])
        append_constraint(rows, lower, upper, {x_index: 1.0, b_index: -float(big_m)}, -np.inf, 0.0)
        append_constraint(rows, lower, upper, {b_index: 1.0, x_index: -1.0}, -np.inf, 0.0)

        # Every worker has exactly one low-rate hour in the shift.
        coefficients = {int(index): 1.0 for index in IDX.low_hour[position]}
        coefficients[x_index] = -1.0
        append_constraint(rows, lower, upper, coefficients, 0.0, 0.0)

    # Hourly processing capacity:
    # 25 * active workers - 15 * workers whose low-rate hour is now.
    for hour in range(HOURS):
        coefficients = {
            int(IDX.processed_early[hour]): 1.0,
            int(IDX.processed_late[hour]): 1.0,
        }
        for position, start in enumerate(STARTS):
            if COVERAGE[hour, position]:
                local_hour = hour - start
                coefficients[int(IDX.x[position])] = -float(NORMAL_RATE)
                coefficients[int(IDX.low_hour[position, local_hour])] = float(RATE_LOSS)
        append_constraint(rows, lower, upper, coefficients, -np.inf, 0.0)

    # Early freight inventory. It arrives in hours 0..11 and must be zero at
    # the end of hour 15 (before 16:00).
    for hour in range(HOURS):
        coefficients = {
            int(IDX.inventory_early[hour]): 1.0,
            int(IDX.processed_early[hour]): 1.0,
        }
        if hour > 0:
            coefficients[int(IDX.inventory_early[hour - 1])] = -1.0
        early_arrival = float(arrivals[hour]) if hour < EARLY_END else 0.0
        append_constraint(rows, lower, upper, coefficients, early_arrival, early_arrival)

    # Late freight inventory. It arrives from hour 12 onward and must be zero
    # by the end of the day.
    for hour in range(HOURS):
        coefficients = {
            int(IDX.inventory_late[hour]): 1.0,
            int(IDX.processed_late[hour]): 1.0,
        }
        if hour > 0:
            coefficients[int(IDX.inventory_late[hour - 1])] = -1.0
        late_arrival = float(arrivals[hour]) if hour >= EARLY_END else 0.0
        append_constraint(rows, lower, upper, coefficients, late_arrival, late_arrival)

    append_constraint(
        rows,
        lower,
        upper,
        {int(IDX.inventory_early[EARLY_DEADLINE - 1]): 1.0},
        0.0,
        0.0,
    )
    append_constraint(rows, lower, upper, {int(IDX.inventory_early[-1]): 1.0}, 0.0, 0.0)
    append_constraint(rows, lower, upper, {int(IDX.inventory_late[-1]): 1.0}, 0.0, 0.0)

    for hour in range(HOURS):
        coefficients = {IDX.peak: -1.0}
        for position in range(len(STARTS)):
            if COVERAGE[hour, position]:
                coefficients[int(IDX.x[position])] = 1.0
        append_constraint(rows, lower, upper, coefficients, -np.inf, 0.0)

    variable_lower = np.zeros(IDX.size, dtype=float)
    variable_upper = np.full(IDX.size, np.inf, dtype=float)
    variable_upper[IDX.b] = 1.0
    variable_upper[IDX.x] = float(big_m)
    variable_upper[IDX.low_hour.ravel()] = float(big_m)
    bounds = Bounds(variable_lower, variable_upper)

    integrality = np.zeros(IDX.size, dtype=int)
    integrality[IDX.b] = 1
    integrality[IDX.x] = 1
    integrality[IDX.low_hour.ravel()] = 1
    integrality[IDX.deviation] = 1
    integrality[IDX.peak] = 1
    return rows, lower, upper, bounds, integrality


def run_milp(
    objective: np.ndarray,
    rows: list[np.ndarray],
    lower: list[float],
    upper: list[float],
    bounds: Bounds,
    integrality: np.ndarray,
    time_limit: float,
    stage: str,
    day: int,
):
    constraints = LinearConstraint(np.vstack(rows), np.array(lower), np.array(upper))
    result = milp(
        c=objective,
        integrality=integrality,
        bounds=bounds,
        constraints=constraints,
        options={
            "disp": False,
            "presolve": True,
            "time_limit": time_limit,
            "mip_rel_gap": 0.0,
        },
    )
    if result.status != 0 or result.x is None:
        raise RuntimeError(
            f"Day {day}, {stage} did not reach an optimum. "
            f"status={result.status}, message={result.message}"
        )
    return result


def solve_day(
    day: int,
    arrivals: np.ndarray,
    time_limit: float,
    fixed_workers: int | None = None,
) -> DaySolution:
    rows, lower, upper, bounds, integrality = build_base_model(arrivals)

    if fixed_workers is None:
        objective = np.zeros(IDX.size)
        objective[IDX.x] = 1.0
        stage1 = run_milp(objective, rows, lower, upper, bounds, integrality, time_limit, "stage 1", day)
        optimum_workers = int(round(float(stage1.x[IDX.x].sum())))
    else:
        if fixed_workers < 5:
            raise ValueError("A fixed-worker schedule needs at least one worker in each of the five shifts.")
        optimum_workers = int(fixed_workers)

    append_constraint(
        rows,
        lower,
        upper,
        {int(index): 1.0 for index in IDX.x},
        optimum_workers,
        optimum_workers,
    )

    objective = np.zeros(IDX.size)
    objective[IDX.peak] = 1.0
    stage2 = run_milp(objective, rows, lower, upper, bounds, integrality, time_limit, "stage 2", day)
    optimum_peak = int(round(float(stage2.x[IDX.peak])))
    append_constraint(rows, lower, upper, {IDX.peak: 1.0}, optimum_peak, optimum_peak)

    # deviation[h] = |3*n[h] - N| at the optimum. This is three times the
    # absolute deviation from average hourly staffing N/3.
    for hour in range(HOURS):
        above_average = {int(IDX.deviation[hour]): 1.0}
        below_average = {int(IDX.deviation[hour]): 1.0}
        for position in range(len(STARTS)):
            if COVERAGE[hour, position]:
                above_average[int(IDX.x[position])] = -3.0
                below_average[int(IDX.x[position])] = 3.0
        append_constraint(rows, lower, upper, above_average, -float(optimum_workers), np.inf)
        append_constraint(rows, lower, upper, below_average, float(optimum_workers), np.inf)

    objective = np.zeros(IDX.size)
    objective[IDX.deviation] = 1.0
    stage3 = run_milp(objective, rows, lower, upper, bounds, integrality, time_limit, "stage 3", day)
    optimum_deviation_scaled = int(round(float(stage3.x[IDX.deviation].sum())))
    append_constraint(
        rows,
        lower,
        upper,
        {int(index): 1.0 for index in IDX.deviation},
        optimum_deviation_scaled,
        optimum_deviation_scaled,
    )

    objective = np.zeros(IDX.size)
    objective[IDX.inventory_early] = 1.0
    objective[IDX.inventory_late] = 1.0
    stage4 = run_milp(objective, rows, lower, upper, bounds, integrality, time_limit, "stage 4", day)

    values = stage4.x
    selected = np.rint(values[IDX.b]).astype(int)
    workers = np.rint(values[IDX.x]).astype(int)
    low_hour_workers = np.rint(values[IDX.low_hour]).astype(int)
    processed_early = values[IDX.processed_early].copy()
    processed_late = values[IDX.processed_late].copy()
    inventory_early = values[IDX.inventory_early].copy()
    inventory_late = values[IDX.inventory_late].copy()
    hourly_workers = np.rint(COVERAGE @ workers).astype(int)

    hourly_low_workers = np.zeros(HOURS, dtype=int)
    for hour in range(HOURS):
        for position, start in enumerate(STARTS):
            if COVERAGE[hour, position]:
                hourly_low_workers[hour] += low_hour_workers[position, hour - start]
    hourly_capacity = NORMAL_RATE * hourly_workers - RATE_LOSS * hourly_low_workers

    solution = DaySolution(
        day=day,
        arrivals=arrivals,
        selected=selected,
        workers_by_start=workers,
        low_hour_workers=low_hour_workers,
        processed_early=processed_early,
        processed_late=processed_late,
        inventory_early=inventory_early,
        inventory_late=inventory_late,
        hourly_workers=hourly_workers,
        hourly_low_workers=hourly_low_workers,
        hourly_capacity=hourly_capacity,
        total_workers=int(workers.sum()),
        peak_workers=int(hourly_workers.max()),
        balance_deviation=float(np.abs(3 * hourly_workers - optimum_workers).sum()) / 3.0,
        theoretical_lower_bound=math.ceil(float(arrivals.sum()) / SHIFT_CAPACITY),
    )
    validate_solution(solution)
    return solution


def validate_solution(solution: DaySolution, tolerance: float = 1e-5) -> None:
    if int(solution.selected.sum()) != 5:
        raise AssertionError(f"Day {solution.day}: expected 5 selected shifts.")
    if np.any((solution.selected == 0) & (solution.workers_by_start != 0)):
        raise AssertionError(f"Day {solution.day}: workers assigned to an unselected shift.")
    if np.any((solution.selected == 1) & (solution.workers_by_start < 1)):
        raise AssertionError(f"Day {solution.day}: selected shift without workers.")
    if not np.array_equal(solution.low_hour_workers.sum(axis=1), solution.workers_by_start):
        raise AssertionError(f"Day {solution.day}: a worker does not have exactly one low-rate hour.")
    if np.any(solution.processed_early < -tolerance) or np.any(solution.processed_late < -tolerance):
        raise AssertionError(f"Day {solution.day}: negative processing amount.")
    total_processed = solution.processed_early + solution.processed_late
    if np.any(total_processed - solution.hourly_capacity > tolerance):
        raise AssertionError(f"Day {solution.day}: hourly capacity exceeded.")

    reconstructed_early = np.empty(HOURS)
    reconstructed_late = np.empty(HOURS)
    previous_early = 0.0
    previous_late = 0.0
    for hour in range(HOURS):
        early_arrival = solution.arrivals[hour] if hour < EARLY_END else 0.0
        late_arrival = solution.arrivals[hour] if hour >= EARLY_END else 0.0
        reconstructed_early[hour] = previous_early + early_arrival - solution.processed_early[hour]
        reconstructed_late[hour] = previous_late + late_arrival - solution.processed_late[hour]
        previous_early = reconstructed_early[hour]
        previous_late = reconstructed_late[hour]

    if not np.allclose(reconstructed_early, solution.inventory_early, atol=tolerance):
        raise AssertionError(f"Day {solution.day}: early inventory balance failed.")
    if not np.allclose(reconstructed_late, solution.inventory_late, atol=tolerance):
        raise AssertionError(f"Day {solution.day}: late inventory balance failed.")
    if abs(solution.inventory_early[EARLY_DEADLINE - 1]) > tolerance:
        raise AssertionError(f"Day {solution.day}: early freight missed the 16:00 deadline.")
    if abs(solution.inventory_early[-1]) > tolerance or abs(solution.inventory_late[-1]) > tolerance:
        raise AssertionError(f"Day {solution.day}: end-of-day freight remains.")
    if solution.peak_workers != int(solution.hourly_workers.max()):
        raise AssertionError(f"Day {solution.day}: inconsistent peak worker count.")

    expected_total_capacity = SHIFT_CAPACITY * solution.total_workers
    if int(solution.hourly_capacity.sum()) != expected_total_capacity:
        raise AssertionError(f"Day {solution.day}: low-rate-hour capacity accounting failed.")


def display_number(value: float):
    rounded = round(float(value))
    if abs(float(value) - rounded) < 1e-6:
        return int(rounded)
    return round(float(value), 4)


def style_worksheet(worksheet, widths: dict[int, float]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = width


def write_results(path: Path, solutions: list[DaySolution]) -> None:
    workbook = Workbook()

    summary = workbook.active
    summary.title = "每日汇总"
    summary.append(
        [
            "天",
            "全天货量",
            "0-12点货量",
            "185件理论下界",
            "最优总人数",
            "峰值在岗人数",
            "平均在岗人数",
            "小时人数总绝对偏差",
            "下界差值",
            "16点早期剩余",
            "日末总剩余",
            "选中班次",
        ]
    )

    shifts = workbook.create_sheet("班次安排")
    shifts.append(["天", "班次序号", "开始时间", "结束时间", "工人数", "班次总处理能力"])

    low_hours = workbook.create_sheet("10件小时安排")
    low_hours.append(["天", "班次开始", "班次内小时", "全局小时", "该小时按10件工作的工人数"])

    hourly = workbook.create_sheet("逐小时详情")
    hourly.append(
        [
            "天",
            "小时",
            "进货量",
            "早期货物处理量",
            "普通货物处理量",
            "总处理量",
            "早期剩余",
            "普通剩余",
            "总剩余",
            "在岗人数",
            "其中10件人数",
            "其中25件人数",
            "最大处理能力",
            "未使用产能",
        ]
    )

    matrix = workbook.create_sheet("班次人数矩阵")
    matrix.append(["天"] + [f"{start}:00-{start + SHIFT_LENGTH}:00" for start in STARTS] + ["总人数"])

    for solution in solutions:
        selected_positions = np.flatnonzero(solution.selected)
        shift_labels = [f"{STARTS[pos]:02d}:00-{STARTS[pos] + SHIFT_LENGTH:02d}:00" for pos in selected_positions]
        summary.append(
            [
                solution.day,
                display_number(solution.arrivals.sum()),
                display_number(solution.arrivals[:EARLY_END].sum()),
                solution.theoretical_lower_bound,
                solution.total_workers,
                solution.peak_workers,
                round(solution.total_workers / 3.0, 4),
                round(solution.balance_deviation, 4),
                solution.total_workers - solution.theoretical_lower_bound,
                display_number(solution.inventory_early[EARLY_DEADLINE - 1]),
                display_number(solution.inventory_early[-1] + solution.inventory_late[-1]),
                "；".join(shift_labels),
            ]
        )

        for order, position in enumerate(selected_positions, start=1):
            start = STARTS[position]
            workers = int(solution.workers_by_start[position])
            shifts.append([solution.day, order, start, start + SHIFT_LENGTH, workers, SHIFT_CAPACITY * workers])
            for local_hour in range(SHIFT_LENGTH):
                low_hours.append(
                    [
                        solution.day,
                        start,
                        local_hour + 1,
                        start + local_hour,
                        int(solution.low_hour_workers[position, local_hour]),
                    ]
                )

        for hour in range(HOURS):
            total_processed = solution.processed_early[hour] + solution.processed_late[hour]
            total_inventory = solution.inventory_early[hour] + solution.inventory_late[hour]
            hourly.append(
                [
                    solution.day,
                    hour,
                    display_number(solution.arrivals[hour]),
                    display_number(solution.processed_early[hour]),
                    display_number(solution.processed_late[hour]),
                    display_number(total_processed),
                    display_number(solution.inventory_early[hour]),
                    display_number(solution.inventory_late[hour]),
                    display_number(total_inventory),
                    int(solution.hourly_workers[hour]),
                    int(solution.hourly_low_workers[hour]),
                    int(solution.hourly_workers[hour] - solution.hourly_low_workers[hour]),
                    int(solution.hourly_capacity[hour]),
                    display_number(solution.hourly_capacity[hour] - total_processed),
                ]
            )

        matrix.append([solution.day] + [int(value) for value in solution.workers_by_start] + [solution.total_workers])

    style_worksheet(summary, {1: 8, 2: 13, 3: 14, 4: 16, 5: 13, 6: 15, 7: 14, 8: 20, 9: 11, 10: 15, 11: 13, 12: 55})
    style_worksheet(shifts, {1: 8, 2: 12, 3: 12, 4: 12, 5: 12, 6: 18})
    style_worksheet(low_hours, {1: 8, 2: 12, 3: 14, 4: 12, 5: 24})
    style_worksheet(hourly, {1: 8, 2: 9, 3: 12, 4: 17, 5: 17, 6: 12, 7: 12, 8: 12, 9: 12, 10: 11, 11: 14, 12: 14, 13: 16, 14: 14})
    style_worksheet(matrix, {1: 8, **{i: 15 for i in range(2, 19)}, 19: 12})

    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Solve Question 2 using a mixed-integer linear program.")
    parser.add_argument("--input", type=Path, default=Path("第一轮B题附件.xlsx"), help="Input Excel file")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("问题二优化结果.xlsx"),
        help="Output Excel file",
    )
    parser.add_argument("--days", type=int, nargs="*", help="Optional day numbers; default is all days")
    parser.add_argument(
        "--time-limit",
        type=float,
        default=120.0,
        help="MILP time limit in seconds for each optimization stage",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    arrivals_by_day = load_arrivals(args.input)
    selected_days = sorted(args.days) if args.days else sorted(arrivals_by_day)
    unknown_days = [day for day in selected_days if day not in arrivals_by_day]
    if unknown_days:
        raise ValueError(f"The input file does not contain days: {unknown_days}")

    solutions: list[DaySolution] = []
    for day in selected_days:
        solution = solve_day(day, arrivals_by_day[day], args.time_limit)
        solutions.append(solution)
        starts = [STARTS[pos] for pos in np.flatnonzero(solution.selected)]
        print(
            f"day={day:02d} workers={solution.total_workers} "
            f"lower={solution.theoretical_lower_bound} peak={solution.peak_workers} "
            f"balance={solution.balance_deviation:.2f} starts={starts}"
        )

    write_results(args.output, solutions)
    print(f"Saved {len(solutions)} day(s) to: {args.output.resolve()}")


if __name__ == "__main__":
    main()
