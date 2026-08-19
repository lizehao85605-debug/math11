"""Independent consistency audit for the Question 3 result workbook."""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from solve_question2 import (
    IDX,
    SHIFT_CAPACITY,
    append_constraint,
    build_base_model,
    load_arrivals,
    run_milp,
    solve_day,
)


NORMAL_RATE = 25
LOW_RATE = 10
RATE_LOSS = NORMAL_RATE - LOW_RATE
TOLERANCE = 1e-4


def close(left: float, right: float, message: str) -> None:
    if abs(float(left) - float(right)) > TOLERANCE:
        raise AssertionError(f"{message}: {left} != {right}")


def maximum_consecutive(flags: list[bool]) -> int:
    result = 0
    current = 0
    for flag in flags:
        current = current + 1 if flag else 0
        result = max(result, current)
    return result


def solve_peak_lower_bound(
    day: int,
    arrivals: np.ndarray,
    worker_cap: int,
    time_limit: float,
) -> int:
    """Re-solve one day without using any schedule stored in the result file."""
    rows, lower, upper, bounds, integrality = build_base_model(arrivals)
    append_constraint(
        rows,
        lower,
        upper,
        {int(index): 1.0 for index in IDX.x},
        -np.inf,
        float(worker_cap),
    )
    objective = np.zeros(IDX.size)
    objective[IDX.peak] = 1.0
    result = run_milp(
        objective,
        rows,
        lower,
        upper,
        bounds,
        integrality,
        time_limit,
        "deep-audit peak lower bound",
        day,
    )
    return int(round(float(result.fun)))


def audit(input_path: Path, result_path: Path, time_limit: float, deep: bool = False) -> None:
    arrivals = load_arrivals(input_path)
    workbook = load_workbook(result_path, data_only=True, read_only=True)
    required_sheets = [
        "招聘汇总",
        "每日人员与检验",
        "第三问班次",
        "10件小时安排",
        "逐小时处理检验",
        "人员检验",
        "人员月历",
        "逐人上班明细",
    ]
    if workbook.sheetnames != required_sheets:
        raise AssertionError(f"Unexpected sheets: {workbook.sheetnames}")

    summary_rows = list(workbook["招聘汇总"].iter_rows(min_row=2, values_only=True))
    summary = {str(row[0]): row[1] for row in summary_rows if row[0] is not None}
    hired = int(summary["最少招聘人数"])
    if hired <= 0:
        raise AssertionError("Hired worker count must be positive.")

    daily_rows = list(workbook["每日人员与检验"].iter_rows(min_row=2, values_only=True))
    daily = {int(row[0]): row for row in daily_rows if row[0] is not None}
    if sorted(daily) != sorted(arrivals):
        raise AssertionError("Daily result does not cover the same days as the input.")

    daily_minimums = {day: int(row[1]) for day, row in daily.items()}
    daily_attendance = {day: int(row[2]) for day, row in daily.items()}
    daily_surplus = {
        day: daily_attendance[day] - daily_minimums[day]
        for day in sorted(daily)
    }
    for day, row in daily.items():
        if int(row[3]) != daily_surplus[day] or daily_surplus[day] < 0:
            raise AssertionError(f"Day {day}: reported staffing surplus is inconsistent.")

    required_worker_days = hired * 23
    if sum(daily_attendance.values()) != required_worker_days:
        raise AssertionError("Daily attendance does not equal hired workers times 23 days.")
    max_surplus = max(daily_surplus.values())
    counting_lower_bound = 0
    while sum(
        min(hired, minimum + counting_lower_bound)
        for minimum in daily_minimums.values()
    ) < required_worker_days:
        counting_lower_bound += 1
    if max_surplus != counting_lower_bound:
        raise AssertionError(
            f"Maximum surplus {max_surplus} does not attain the counting lower bound "
            f"{counting_lower_bound}."
        )

    shift_rows = list(workbook["第三问班次"].iter_rows(min_row=2, values_only=True))
    shifts: dict[int, list[tuple[int, int, int]]] = defaultdict(list)
    for row in shift_rows:
        day, _, start, end, workers, reported_capacity = row
        day, start, end, workers = int(day), int(start), int(end), int(workers)
        if not 0 <= start <= 16 or end - start != 8:
            raise AssertionError(f"Day {day}: invalid shift {start}-{end}.")
        if workers < 1:
            raise AssertionError(f"Day {day}: selected shift has no workers.")
        close(reported_capacity, SHIFT_CAPACITY * workers, f"Day {day}, shift capacity")
        shifts[day].append((start, end, workers))

    for day in sorted(arrivals):
        if len(shifts[day]) != 5:
            raise AssertionError(f"Day {day}: expected 5 shifts, found {len(shifts[day])}.")
        if len({start for start, _, _ in shifts[day]}) != 5:
            raise AssertionError(f"Day {day}: shift starts are not distinct.")
        actual_workers = int(daily[day][2])
        if sum(workers for _, _, workers in shifts[day]) != actual_workers:
            raise AssertionError(f"Day {day}: shift workers do not equal actual attendance.")
        if int(daily[day][4]) != hired - actual_workers:
            raise AssertionError(f"Day {day}: rest-worker count is inconsistent.")

    low_rows = list(workbook["10件小时安排"].iter_rows(min_row=2, values_only=True))
    low_by_shift: dict[tuple[int, int], dict[int, int]] = defaultdict(dict)
    low_by_global: dict[tuple[int, int], int] = defaultdict(int)
    for row in low_rows:
        day, start, local_hour, global_hour, workers = map(int, row)
        if not 1 <= local_hour <= 8 or global_hour != start + local_hour - 1:
            raise AssertionError(f"Day {day}, shift {start}: invalid low-rate hour mapping.")
        if workers < 0:
            raise AssertionError(f"Day {day}, shift {start}: negative low-rate workers.")
        if local_hour in low_by_shift[(day, start)]:
            raise AssertionError(f"Day {day}, shift {start}: duplicate local hour {local_hour}.")
        low_by_shift[(day, start)][local_hour] = workers
        low_by_global[(day, global_hour)] += workers

    for day in sorted(arrivals):
        for start, _, workers in shifts[day]:
            local = low_by_shift[(day, start)]
            if sorted(local) != list(range(1, 9)):
                raise AssertionError(f"Day {day}, shift {start}: low-rate schedule is incomplete.")
            if sum(local.values()) != workers:
                raise AssertionError(
                    f"Day {day}, shift {start}: not every worker has exactly one 10-item hour."
                )

    hourly_rows = list(workbook["逐小时处理检验"].iter_rows(min_row=2, values_only=True))
    hourly: dict[tuple[int, int], tuple] = {}
    for row in hourly_rows:
        key = (int(row[0]), int(row[1]))
        if key in hourly:
            raise AssertionError(f"Duplicate hourly row: {key}")
        hourly[key] = row

    total_unused_capacity = 0.0
    fractional_hourly_cells: list[tuple[int, int, int, float]] = []
    for day in sorted(arrivals):
        previous_early = 0.0
        previous_late = 0.0
        total_processed = 0.0
        for hour in range(24):
            row = hourly[(day, hour)]
            (
                _,
                _,
                reported_arrival,
                processed_early,
                processed_late,
                processed_total,
                inventory_early,
                inventory_late,
                inventory_total,
                on_duty,
                low_workers,
                capacity,
                unused_capacity,
            ) = row
            close(reported_arrival, arrivals[day][hour], f"Day {day}, hour {hour}, arrival")
            if min(processed_early, processed_late, inventory_early, inventory_late) < -TOLERANCE:
                raise AssertionError(f"Day {day}, hour {hour}: negative processing or inventory.")
            close(processed_total, processed_early + processed_late, f"Day {day}, hour {hour}, processing")
            close(inventory_total, inventory_early + inventory_late, f"Day {day}, hour {hour}, inventory")

            expected_on_duty = sum(
                workers for start, end, workers in shifts[day] if start <= hour < end
            )
            if int(on_duty) != expected_on_duty:
                raise AssertionError(f"Day {day}, hour {hour}: on-duty count mismatch.")
            expected_low = low_by_global[(day, hour)]
            if int(low_workers) != expected_low:
                raise AssertionError(f"Day {day}, hour {hour}: low-rate worker count mismatch.")
            expected_capacity = NORMAL_RATE * expected_on_duty - RATE_LOSS * expected_low
            close(capacity, expected_capacity, f"Day {day}, hour {hour}, capacity")
            if processed_total - capacity > TOLERANCE:
                raise AssertionError(f"Day {day}, hour {hour}: processing exceeds capacity.")
            close(unused_capacity, capacity - processed_total, f"Day {day}, hour {hour}, unused capacity")

            early_arrival = arrivals[day][hour] if hour < 12 else 0.0
            late_arrival = arrivals[day][hour] if hour >= 12 else 0.0
            expected_early_inventory = previous_early + early_arrival - processed_early
            expected_late_inventory = previous_late + late_arrival - processed_late
            close(inventory_early, expected_early_inventory, f"Day {day}, hour {hour}, early balance")
            close(inventory_late, expected_late_inventory, f"Day {day}, hour {hour}, late balance")
            previous_early = float(inventory_early)
            previous_late = float(inventory_late)
            total_processed += float(processed_total)
            total_unused_capacity += float(unused_capacity)
            for column, value in enumerate(row[2:13], start=3):
                if (
                    isinstance(value, (int, float))
                    and abs(float(value) - round(float(value))) > TOLERANCE
                ):
                    fractional_hourly_cells.append((day, hour, column, float(value)))

        close(hourly[(day, 15)][6], 0.0, f"Day {day}, 16:00 early deadline")
        close(previous_early, 0.0, f"Day {day}, final early inventory")
        close(previous_late, 0.0, f"Day {day}, final late inventory")
        close(total_processed, arrivals[day].sum(), f"Day {day}, total processing")
        close(daily[day][5], 0.0, f"Day {day}, reported deadline inventory")
        close(daily[day][6], 0.0, f"Day {day}, reported final inventory")

    if fractional_hourly_cells:
        raise AssertionError(
            "Hourly freight results contain fractional item counts: "
            f"{fractional_hourly_cells[:5]}"
        )

    calendar_rows = list(workbook["人员月历"].iter_rows(min_row=2, values_only=True))
    if len(calendar_rows) != hired:
        raise AssertionError(f"Expected {hired} worker calendars, found {len(calendar_rows)}.")
    calendar_assignments: set[tuple[int, int, str]] = set()
    assignment_counts: Counter[tuple[int, int]] = Counter()
    max_streak_seen = 0
    for row in calendar_rows:
        worker = int(row[0])
        entries = list(row[1:31])
        flags = [entry != "休" for entry in entries]
        worked = sum(flags)
        max_streak = maximum_consecutive(flags)
        if worked != 23 or int(row[31]) != 23:
            raise AssertionError(f"Worker {worker}: does not work exactly 23 days.")
        if max_streak > 7 or int(row[32]) != max_streak:
            raise AssertionError(f"Worker {worker}: consecutive-work check failed.")
        max_streak_seen = max(max_streak_seen, max_streak)

        for day_index, entry in enumerate(entries, start=1):
            if entry == "休":
                continue
            start_text, end_text = str(entry).split("-")
            start, end = int(start_text), int(end_text)
            if (start, end) not in {(s, e) for s, e, _ in shifts[day_index]}:
                raise AssertionError(f"Worker {worker}, day {day_index}: unknown shift {entry}.")
            calendar_assignments.add((worker, day_index, str(entry)))
            assignment_counts[(day_index, start)] += 1

    for day in sorted(arrivals):
        for start, _, workers in shifts[day]:
            if assignment_counts[(day, start)] != workers:
                raise AssertionError(f"Day {day}, shift {start}: named-worker assignment mismatch.")

    detail_rows = list(workbook["逐人上班明细"].iter_rows(min_row=2, values_only=True))
    detail_assignments = {(int(row[0]), int(row[1]), str(row[2])) for row in detail_rows}
    if len(detail_assignments) != len(detail_rows):
        raise AssertionError("Duplicate rows exist in the worker assignment details.")
    if detail_assignments != calendar_assignments:
        raise AssertionError("Worker detail sheet and monthly calendar do not match.")

    has_individual_low_hours = bool(detail_rows) and len(detail_rows[0]) >= 5
    if has_individual_low_hours:
        individual_low_counts: Counter[tuple[int, int, int]] = Counter()
        worker_low_counts: Counter[int] = Counter()
        for row in detail_rows:
            worker, day, shift, global_hour, local_hour = row[:5]
            worker, day = int(worker), int(day)
            global_hour, local_hour = int(global_hour), int(local_hour)
            start_text, end_text = str(shift).split("-")
            start, end = int(start_text), int(end_text)
            if not 1 <= local_hour <= 8 or global_hour != start + local_hour - 1:
                raise AssertionError(
                    f"Worker {worker}, day {day}: invalid individual 10-item hour."
                )
            if not start <= global_hour < end:
                raise AssertionError(
                    f"Worker {worker}, day {day}: 10-item hour is outside the shift."
                )
            individual_low_counts[(day, start, local_hour)] += 1
            worker_low_counts[worker] += 1

        for day in sorted(arrivals):
            for start, _, _ in shifts[day]:
                for local_hour, expected in low_by_shift[(day, start)].items():
                    actual = individual_low_counts[(day, start, local_hour)]
                    if actual != expected:
                        raise AssertionError(
                            f"Day {day}, shift {start}, local hour {local_hour}: "
                            f"individual count {actual} != aggregate count {expected}."
                        )
        if any(worker_low_counts[worker] != 23 for worker in range(1, hired + 1)):
            raise AssertionError("A worker does not have exactly 23 individual 10-item hours.")

    stats_rows = list(workbook["人员检验"].iter_rows(min_row=2, values_only=True))
    if len(stats_rows) != hired or any(row[4] != "是" for row in stats_rows):
        raise AssertionError("Worker validation sheet contains an invalid worker.")

    # Any monthly plan must cover the hardest single day under Question 2.
    # Deep mode re-solves every daily minimum and every daily peak lower bound.
    hardest_day = max(daily_minimums, key=daily_minimums.get)
    if deep:
        resolved_minimums = {
            day: solve_day(day, arrivals[day], time_limit).total_workers
            for day in sorted(arrivals)
        }
        mismatched_minimums = {
            day: (daily_minimums[day], resolved_minimums[day])
            for day in sorted(arrivals)
            if daily_minimums[day] != resolved_minimums[day]
        }
        if mismatched_minimums:
            raise AssertionError(f"Daily minimum worker counts changed: {mismatched_minimums}")
        hardest_minimum = resolved_minimums[hardest_day]
    else:
        hardest_minimum = solve_day(
            hardest_day, arrivals[hardest_day], time_limit
        ).total_workers
        if hardest_minimum != daily_minimums[hardest_day]:
            raise AssertionError("Reported daily lower bound does not match an independent re-solve.")

    if hired != hardest_minimum:
        raise AssertionError(
            f"Hired count {hired} does not meet the proven lower bound {hardest_minimum}."
        )

    actual_daily_peaks = {
        day: max(int(hourly[(day, hour)][9]) for hour in range(24))
        for day in sorted(arrivals)
    }
    if deep:
        peak_lower_bounds = {
            day: solve_peak_lower_bound(day, arrivals[day], hired, time_limit)
            for day in sorted(arrivals)
        }
        mismatched_peaks = {
            day: (actual_daily_peaks[day], peak_lower_bounds[day])
            for day in sorted(arrivals)
            if actual_daily_peaks[day] != peak_lower_bounds[day]
        }
        if mismatched_peaks:
            raise AssertionError(f"Daily peak lower bounds are not attained: {mismatched_peaks}")

    print("PASS: all 30 days contain five distinct 8-hour shifts.")
    print("PASS: every shift worker has exactly one aggregated 10-item hour.")
    print("PASS: all 720 hourly capacity and inventory balances are valid.")
    print("PASS: all early freight is complete before 16:00 and all daily freight is complete by 24:00.")
    print(f"PASS: all {hired} workers work exactly 23 days and at most {max_streak_seen} consecutive days.")
    print(f"PASS: {len(calendar_assignments)} named assignments match all shift totals.")
    if has_individual_low_hours:
        print("PASS: every named assignment has one valid individual 10-item hour.")
    print(
        f"PASS: optimality certified by day {hardest_day}: at least "
        f"{hardest_minimum} workers are necessary and {hired} are sufficient."
    )
    print(
        f"PASS: maximum daily surplus {max_surplus} attains its worker-day counting "
        f"lower bound {counting_lower_bound}."
    )
    print("PASS: all reported hourly item quantities are integral.")
    if deep:
        print("PASS: all 30 daily minimum worker counts were independently re-solved.")
        print(
            "PASS: all 30 daily staffing peaks attain their independent lower bounds; "
            f"global peak={max(actual_daily_peaks.values())}, "
            f"sum of daily peaks={sum(actual_daily_peaks.values())}."
        )
    print(f"INFO: total unused processing capacity across the month = {total_unused_capacity:.2f} items.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Verify the Question 3 result workbook.")
    parser.add_argument("--input", type=Path, default=Path("第一轮B题附件.xlsx"))
    parser.add_argument("--result", type=Path, default=Path("问题三联合优化结果.xlsx"))
    parser.add_argument("--time-limit", type=float, default=180.0)
    parser.add_argument(
        "--deep",
        action="store_true",
        help="Re-solve all 30 daily minima and all 30 daily peak lower bounds.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    audit(args.input, args.result, args.time_limit, args.deep)
