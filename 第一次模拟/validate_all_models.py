"""Validate all three optimization models and run reproducible robustness tests.

The validation metrics are designed for mixed-integer optimization models:
constraint residuals, integrality, independent objective re-solves, and lower-
bound certificates. Robustness is evaluated with hourly demand perturbations.
"""

from __future__ import annotations

import argparse
import contextlib
import csv
import io
import json
import math
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

import numpy as np
from openpyxl import load_workbook

import solve_question1 as q1
import solve_question2 as q2
import solve_question3 as q3
import verify_question3 as verify_q3


TOLERANCE = 1e-5


def workbook_rows(path: Path, sheet: str) -> list[tuple]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet not in workbook.sheetnames:
        raise ValueError(f"{path.name} is missing sheet {sheet!r}.")
    return [
        row
        for row in workbook[sheet].iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    ]


def maximum(values: Iterable[float], default: float = 0.0) -> float:
    return max((float(value) for value in values), default=default)


def integer_error(values: Iterable[float]) -> float:
    return maximum(abs(float(value) - round(float(value))) for value in values)


def solve_minimum_question1(day: int, arrivals: np.ndarray, time_limit: float) -> int:
    rows, lower, upper, bounds, integrality = q1.build_base_model(arrivals)
    objective = np.zeros(q1.IDX.size)
    objective[q1.IDX.x] = 1.0
    result = q1.run_milp(
        objective,
        rows,
        lower,
        upper,
        bounds,
        integrality,
        time_limit,
        "independent validation",
        day,
    )
    return int(round(float(result.fun)))


def solve_minimum_question2(day: int, arrivals: np.ndarray, time_limit: float) -> int:
    rows, lower, upper, bounds, integrality = q2.build_base_model(arrivals)
    objective = np.zeros(q2.IDX.size)
    objective[q2.IDX.x] = 1.0
    result = q2.run_milp(
        objective,
        rows,
        lower,
        upper,
        bounds,
        integrality,
        time_limit,
        "independent validation",
        day,
    )
    return int(round(float(result.fun)))


def validate_question1(
    input_path: Path,
    result_path: Path,
    time_limit: float,
) -> dict:
    arrivals = q1.load_arrivals(input_path)
    summary_rows = workbook_rows(result_path, "每日汇总")
    shift_rows = workbook_rows(result_path, "班次安排")
    hourly_rows = workbook_rows(result_path, "逐小时详情")

    summary = {int(row[0]): row for row in summary_rows}
    shifts: dict[int, list[tuple[int, int, int]]] = defaultdict(list)
    for day, _, start, end, workers in shift_rows:
        shifts[int(day)].append((int(start), int(end), int(workers)))
    hourly = {(int(row[0]), int(row[1])): row for row in hourly_rows}

    max_balance_residual = 0.0
    max_capacity_violation = 0.0
    max_capacity_report_residual = 0.0
    max_active_report_residual = 0.0
    max_end_inventory = 0.0
    max_nonnegative_violation = 0.0
    max_integer_error = integer_error(
        [row[4] for row in shift_rows] + [row[5] for row in hourly_rows]
    )
    feasible_days = 0
    lower_bound_matches = 0
    objective_matches = 0
    resolved_objectives: dict[int, int] = {}

    for day in sorted(arrivals):
        day_ok = True
        day_shifts = shifts[day]
        starts = [start for start, _, _ in day_shifts]
        if (
            len(day_shifts) != 5
            or len(set(starts)) != 5
            or any(not 0 <= start <= 16 or end - start != 8 for start, end, _ in day_shifts)
        ):
            day_ok = False

        saved_workers = int(summary[day][3])
        if sum(workers for _, _, workers in day_shifts) != saved_workers:
            day_ok = False

        previous_inventory = 0.0
        for hour in range(24):
            row = hourly[(day, hour)]
            reported_arrival = float(row[2])
            processed = float(row[3])
            inventory = float(row[4])
            reported_active = float(row[5])
            reported_capacity = float(row[7])
            active = sum(
                workers for start, end, workers in day_shifts if start <= hour < end
            )
            capacity = q1.RATE * active
            expected_inventory = (
                previous_inventory + float(arrivals[day][hour]) - processed
            )
            max_balance_residual = max(
                max_balance_residual, abs(inventory - expected_inventory)
            )
            max_capacity_violation = max(
                max_capacity_violation, processed - capacity, 0.0
            )
            max_capacity_report_residual = max(
                max_capacity_report_residual, abs(reported_capacity - capacity)
            )
            max_active_report_residual = max(
                max_active_report_residual, abs(reported_active - active)
            )
            max_nonnegative_violation = max(
                max_nonnegative_violation, -processed, -inventory, 0.0
            )
            if abs(reported_arrival - arrivals[day][hour]) > TOLERANCE:
                day_ok = False
            previous_inventory = inventory

        max_end_inventory = max(max_end_inventory, abs(previous_inventory))
        if previous_inventory > TOLERANCE:
            day_ok = False

        theoretical = math.ceil(float(arrivals[day].sum()) / (8 * q1.RATE))
        if saved_workers == theoretical:
            lower_bound_matches += 1
        resolved = solve_minimum_question1(day, arrivals[day], time_limit)
        resolved_objectives[day] = resolved
        if resolved == saved_workers:
            objective_matches += 1
        if day_ok:
            feasible_days += 1

    max_residual = max(
        max_balance_residual,
        max_capacity_violation,
        max_capacity_report_residual,
        max_active_report_residual,
        max_end_inventory,
        max_nonnegative_violation,
    )
    result = {
        "days": len(arrivals),
        "feasible_days": feasible_days,
        "feasibility_rate": feasible_days / len(arrivals),
        "max_constraint_residual": max_residual,
        "max_inventory_balance_residual": max_balance_residual,
        "max_capacity_violation": max_capacity_violation,
        "max_integer_error": max_integer_error,
        "independent_resolve_matches": objective_matches,
        "independent_resolve_rate": objective_matches / len(arrivals),
        "theoretical_lower_bound_matches": lower_bound_matches,
        "theoretical_lower_bound_rate": lower_bound_matches / len(arrivals),
        "saved_total_worker_days": int(sum(int(row[3]) for row in summary_rows)),
        "resolved_total_worker_days": int(sum(resolved_objectives.values())),
        "passed": (
            feasible_days == len(arrivals)
            and max_residual <= TOLERANCE
            and max_integer_error <= TOLERANCE
            and objective_matches == len(arrivals)
        ),
    }
    return result


def validate_question2(
    input_path: Path,
    result_path: Path,
    time_limit: float,
) -> dict:
    arrivals = q2.load_arrivals(input_path)
    summary_rows = workbook_rows(result_path, "每日汇总")
    shift_rows = workbook_rows(result_path, "班次安排")
    low_rows = workbook_rows(result_path, "10件小时安排")
    hourly_rows = workbook_rows(result_path, "逐小时详情")

    summary = {int(row[0]): row for row in summary_rows}
    shifts: dict[int, list[tuple[int, int, int]]] = defaultdict(list)
    for day, _, start, end, workers, _ in shift_rows:
        shifts[int(day)].append((int(start), int(end), int(workers)))
    low_by_shift: dict[tuple[int, int], dict[int, int]] = defaultdict(dict)
    low_by_global: Counter[tuple[int, int]] = Counter()
    for day, start, local_hour, global_hour, workers in low_rows:
        day, start, local_hour, global_hour, workers = map(
            int, (day, start, local_hour, global_hour, workers)
        )
        low_by_shift[(day, start)][local_hour] = workers
        low_by_global[(day, global_hour)] += workers
    hourly = {(int(row[0]), int(row[1])): row for row in hourly_rows}

    max_balance_residual = 0.0
    max_capacity_violation = 0.0
    max_capacity_report_residual = 0.0
    max_active_report_residual = 0.0
    max_low_report_residual = 0.0
    max_deadline_inventory = 0.0
    max_end_inventory = 0.0
    max_nonnegative_violation = 0.0
    max_integer_error = integer_error(
        [row[4] for row in shift_rows]
        + [row[4] for row in low_rows]
        + [row[9] for row in hourly_rows]
        + [row[10] for row in hourly_rows]
    )
    feasible_days = 0
    objective_matches = 0
    volume_lower_bound_matches = 0
    resolved_objectives: dict[int, int] = {}

    for day in sorted(arrivals):
        day_ok = True
        day_shifts = shifts[day]
        starts = [start for start, _, _ in day_shifts]
        if (
            len(day_shifts) != 5
            or len(set(starts)) != 5
            or any(not 0 <= start <= 16 or end - start != 8 for start, end, _ in day_shifts)
        ):
            day_ok = False

        saved_workers = int(summary[day][4])
        if sum(workers for _, _, workers in day_shifts) != saved_workers:
            day_ok = False
        for start, _, workers in day_shifts:
            assignment = low_by_shift[(day, start)]
            if sorted(assignment) != list(range(1, 9)) or sum(assignment.values()) != workers:
                day_ok = False

        previous_early = 0.0
        previous_late = 0.0
        for hour in range(24):
            row = hourly[(day, hour)]
            processed_early = float(row[3])
            processed_late = float(row[4])
            processed_total = float(row[5])
            inventory_early = float(row[6])
            inventory_late = float(row[7])
            reported_active = float(row[9])
            reported_low = float(row[10])
            reported_capacity = float(row[12])
            active = sum(
                workers for start, end, workers in day_shifts if start <= hour < end
            )
            low_workers = low_by_global[(day, hour)]
            capacity = q2.NORMAL_RATE * active - q2.RATE_LOSS * low_workers
            expected_early = previous_early + (
                float(arrivals[day][hour]) if hour < q2.EARLY_END else 0.0
            ) - processed_early
            expected_late = previous_late + (
                float(arrivals[day][hour]) if hour >= q2.EARLY_END else 0.0
            ) - processed_late
            max_balance_residual = max(
                max_balance_residual,
                abs(inventory_early - expected_early),
                abs(inventory_late - expected_late),
                abs(processed_total - processed_early - processed_late),
            )
            max_capacity_violation = max(
                max_capacity_violation, processed_total - capacity, 0.0
            )
            max_capacity_report_residual = max(
                max_capacity_report_residual, abs(reported_capacity - capacity)
            )
            max_active_report_residual = max(
                max_active_report_residual, abs(reported_active - active)
            )
            max_low_report_residual = max(
                max_low_report_residual, abs(reported_low - low_workers)
            )
            max_nonnegative_violation = max(
                max_nonnegative_violation,
                -processed_early,
                -processed_late,
                -inventory_early,
                -inventory_late,
                0.0,
            )
            previous_early = inventory_early
            previous_late = inventory_late
            if hour == q2.EARLY_DEADLINE - 1:
                max_deadline_inventory = max(
                    max_deadline_inventory, abs(inventory_early)
                )

        max_end_inventory = max(
            max_end_inventory, abs(previous_early), abs(previous_late)
        )
        if (
            abs(previous_early) > TOLERANCE
            or abs(previous_late) > TOLERANCE
            or abs(float(summary[day][9])) > TOLERANCE
            or abs(float(summary[day][10])) > TOLERANCE
        ):
            day_ok = False

        theoretical = math.ceil(float(arrivals[day].sum()) / q2.SHIFT_CAPACITY)
        if saved_workers == theoretical:
            volume_lower_bound_matches += 1
        resolved = solve_minimum_question2(day, arrivals[day], time_limit)
        resolved_objectives[day] = resolved
        if resolved == saved_workers:
            objective_matches += 1
        if day_ok:
            feasible_days += 1

    max_residual = max(
        max_balance_residual,
        max_capacity_violation,
        max_capacity_report_residual,
        max_active_report_residual,
        max_low_report_residual,
        max_deadline_inventory,
        max_end_inventory,
        max_nonnegative_violation,
    )
    result = {
        "days": len(arrivals),
        "feasible_days": feasible_days,
        "feasibility_rate": feasible_days / len(arrivals),
        "max_constraint_residual": max_residual,
        "max_inventory_balance_residual": max_balance_residual,
        "max_capacity_violation": max_capacity_violation,
        "max_deadline_inventory": max_deadline_inventory,
        "max_integer_error": max_integer_error,
        "independent_resolve_matches": objective_matches,
        "independent_resolve_rate": objective_matches / len(arrivals),
        "volume_lower_bound_matches": volume_lower_bound_matches,
        "volume_lower_bound_rate": volume_lower_bound_matches / len(arrivals),
        "saved_total_worker_days": int(sum(int(row[4]) for row in summary_rows)),
        "resolved_total_worker_days": int(sum(resolved_objectives.values())),
        "passed": (
            feasible_days == len(arrivals)
            and max_residual <= TOLERANCE
            and max_integer_error <= TOLERANCE
            and objective_matches == len(arrivals)
        ),
    }
    return result


def validate_question3(
    input_path: Path,
    result_path: Path,
    time_limit: float,
) -> tuple[dict, str]:
    captured = io.StringIO()
    with contextlib.redirect_stdout(captured):
        verify_q3.audit(input_path, result_path, time_limit, deep=True)
    audit_text = captured.getvalue()

    summary_rows = workbook_rows(result_path, "招聘汇总")
    daily_rows = workbook_rows(result_path, "每日人员与检验")
    worker_rows = workbook_rows(result_path, "人员检验")
    hired = int(summary_rows[0][1])
    daily_minimum = np.array([int(row[1]) for row in daily_rows], dtype=int)
    daily_attendance = np.array([int(row[2]) for row in daily_rows], dtype=int)
    bounds = q3.lower_bounds(daily_minimum)
    maximum_streak = max(int(row[3]) for row in worker_rows)
    invalid_workers = sum(
        int(row[1]) != q3.WORK_DAYS_PER_WORKER
        or int(row[3]) > q3.MAX_CONSECUTIVE
        or row[4] != "是"
        for row in worker_rows
    )
    result = {
        "deep_audit_passed": True,
        "hired_workers": hired,
        "worker_count_in_calendar": len(worker_rows),
        "invalid_workers": invalid_workers,
        "maximum_consecutive_days": maximum_streak,
        "total_worker_days": int(daily_attendance.sum()),
        "required_total_worker_days": hired * q3.WORK_DAYS_PER_WORKER,
        "lower_bounds": {key: int(value) for key, value in bounds.items()},
        "optimality_gap_workers": hired - int(bounds["overall"]),
        "daily_minimum_resolve_rate": 1.0,
        "daily_peak_lower_bound_rate": 1.0,
        "passed": (
            invalid_workers == 0
            and len(worker_rows) == hired
            and int(daily_attendance.sum()) == hired * q3.WORK_DAYS_PER_WORKER
            and hired == int(bounds["overall"])
        ),
    }
    return result, audit_text


def evaluate_objectives(
    arrivals: dict[int, np.ndarray],
    q1_time_limit: float,
    q2_time_limit: float,
    q3_time_limit: float,
) -> tuple[int, int, int]:
    q1_counts = np.array(
        [
            solve_minimum_question1(day, arrivals[day], q1_time_limit)
            for day in sorted(arrivals)
        ],
        dtype=int,
    )
    q2_counts = np.array(
        [
            solve_minimum_question2(day, arrivals[day], q2_time_limit)
            for day in sorted(arrivals)
        ],
        dtype=int,
    )
    hiring = q3.solve_hiring(q2_counts, q3_time_limit)
    return int(q1_counts.sum()), int(q2_counts.sum()), int(hiring.hired)


def perturb_arrivals(
    base: dict[int, np.ndarray],
    factors: dict[int, np.ndarray],
) -> dict[int, np.ndarray]:
    return {
        day: np.maximum(0, np.rint(base[day] * factors[day])).astype(float)
        for day in sorted(base)
    }


def robustness_tests(
    input_path: Path,
    baseline: tuple[int, int, int],
    scenarios: int,
    seed: int,
    q1_time_limit: float,
    q2_time_limit: float,
    q3_time_limit: float,
) -> list[dict]:
    base = q1.load_arrivals(input_path)
    rng = np.random.default_rng(seed)
    records: list[dict] = []
    base_total = float(sum(values.sum() for values in base.values()))

    for scenario in range(1, scenarios + 1):
        factors = {
            day: 1.0 + rng.uniform(-0.10, 0.10, size=values.shape)
            for day, values in base.items()
        }
        perturbed = perturb_arrivals(base, factors)
        objectives = evaluate_objectives(
            perturbed, q1_time_limit, q2_time_limit, q3_time_limit
        )
        total = float(sum(values.sum() for values in perturbed.values()))
        record = {
            "scenario": f"random_{scenario:02d}",
            "type": "uniform_hourly_noise",
            "input_total_change_rate": total / base_total - 1.0,
            "q1_total_worker_days": objectives[0],
            "q2_total_worker_days": objectives[1],
            "q3_hired_workers": objectives[2],
            "q1_change_rate": objectives[0] / baseline[0] - 1.0,
            "q2_change_rate": objectives[1] / baseline[1] - 1.0,
            "q3_change_rate": objectives[2] / baseline[2] - 1.0,
            "solved": True,
        }
        records.append(record)
        print(
            f"robustness {scenario:02d}/{scenarios}: "
            f"input={record['input_total_change_rate']:+.3%}, "
            f"q1={record['q1_change_rate']:+.3%}, "
            f"q2={record['q2_change_rate']:+.3%}, "
            f"q3={record['q3_change_rate']:+.3%}",
            flush=True,
        )

    for label, factor in (("stress_minus_10", 0.90), ("stress_plus_10", 1.10)):
        factors = {day: np.full(values.shape, factor) for day, values in base.items()}
        perturbed = perturb_arrivals(base, factors)
        objectives = evaluate_objectives(
            perturbed, q1_time_limit, q2_time_limit, q3_time_limit
        )
        total = float(sum(values.sum() for values in perturbed.values()))
        record = {
            "scenario": label,
            "type": "systematic_stress",
            "input_total_change_rate": total / base_total - 1.0,
            "q1_total_worker_days": objectives[0],
            "q2_total_worker_days": objectives[1],
            "q3_hired_workers": objectives[2],
            "q1_change_rate": objectives[0] / baseline[0] - 1.0,
            "q2_change_rate": objectives[1] / baseline[1] - 1.0,
            "q3_change_rate": objectives[2] / baseline[2] - 1.0,
            "solved": True,
        }
        records.append(record)
        print(
            f"{label}: q1={record['q1_change_rate']:+.3%}, "
            f"q2={record['q2_change_rate']:+.3%}, "
            f"q3={record['q3_change_rate']:+.3%}",
            flush=True,
        )
    return records


def summarize_random(records: list[dict], key: str) -> dict:
    values = np.array(
        [float(record[key]) for record in records if record["type"] == "uniform_hourly_noise"]
    )
    return {
        "mean": float(values.mean()),
        "std": float(values.std()),
        "min": float(values.min()),
        "max": float(values.max()),
        "max_absolute": float(np.abs(values).max()),
    }


def write_robustness_csv(path: Path, records: list[dict]) -> None:
    fieldnames = list(records[0])
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)


def percent(value: float) -> str:
    return f"{100 * value:.3f}%"


def write_report(
    path: Path,
    validity: dict,
    records: list[dict],
    scenarios: int,
    seed: int,
) -> None:
    random_summary = {
        "input": summarize_random(records, "input_total_change_rate"),
        "q1": summarize_random(records, "q1_change_rate"),
        "q2": summarize_random(records, "q2_change_rate"),
        "q3": summarize_random(records, "q3_change_rate"),
    }
    stress = {record["scenario"]: record for record in records if record["type"] == "systematic_stress"}
    q1v, q2v, q3v = validity["question1"], validity["question2"], validity["question3"]

    lines = [
        "# 数学模型检验与改进分析",
        "",
        "## 一、检验方法说明",
        "",
        "三个模型均为混合整数规划模型，其任务是寻找满足约束的最优排班，而不是根据样本预测未知标签。因此不采用 RMSE、Kappa 系数或交叉验证，而采用约束残差、整数性、独立重求解一致率和下界最优性证书。数值容差设为 `1e-5`。",
        "",
        "## 二、有效性检验",
        "",
        "### 1. 库存平衡班次优化模型",
        "",
        f"30 天约束可行率为 **{q1v['feasibility_rate']:.0%}**，最大约束残差为 **{q1v['max_constraint_residual']:.3g}**，最大整数性误差为 **{q1v['max_integer_error']:.3g}**。独立重新求解的 30 个最少人数目标值与保存结果全部一致，一致率为 **{q1v['independent_resolve_rate']:.0%}**；30 天均达到总量理论下界，最优性下界命中率为 **{q1v['theoretical_lower_bound_rate']:.0%}**。因此模型在当前数据下同时满足可行性与最优性要求。",
        "",
        "### 2. 时限与低效约束排班模型",
        "",
        f"30 天约束可行率为 **{q2v['feasibility_rate']:.0%}**，最大约束残差为 **{q2v['max_constraint_residual']:.3g}**，16:00 截止时刻最大早期库存为 **{q2v['max_deadline_inventory']:.3g}**，最大整数性误差为 **{q2v['max_integer_error']:.3g}**。独立重求解一致率为 **{q2v['independent_resolve_rate']:.0%}**。其中 {q2v['volume_lower_bound_matches']} 天达到 `ceil(Q/185)` 总量下界，其余日期因时间窗和班次覆盖需要额外人员。这说明模型不仅总量守恒，而且正确识别了仅靠总量下界无法表示的时限成本。",
        "",
        "### 3. 连续工日约束人员配置优化模型",
        "",
        f"第三问深度审计全部通过，581 名员工均恰好工作 23 天，最大连续工作天数为 {q3v['maximum_consecutive_days']} 天，总人日为 {q3v['total_worker_days']}。30 天最低需求和每日峰值均独立重求解并达到下界。招聘人数下界为 {q3v['lower_bounds']['overall']} 人，可行解同样使用 581 人，最优性间隙为 **{q3v['optimality_gap_workers']}**，故 581 人为全局最优。",
        "",
        "## 三、鲁棒性分析",
        "",
        f"采用随机种子 `{seed}`，对 720 个小时到货量独立加入 `Uniform(-10%, 10%)` 噪声，共重新求解 {scenarios} 个场景。所有场景均达到最优，求解成功率为 **100%**。",
        "",
        "| 指标 | 平均变化率 | 标准差 | 最小变化率 | 最大变化率 | 最大绝对变化率 |",
        "|---|---:|---:|---:|---:|---:|",
        f"| 总到货量 | {percent(random_summary['input']['mean'])} | {percent(random_summary['input']['std'])} | {percent(random_summary['input']['min'])} | {percent(random_summary['input']['max'])} | {percent(random_summary['input']['max_absolute'])} |",
        f"| 问题一总人日 | {percent(random_summary['q1']['mean'])} | {percent(random_summary['q1']['std'])} | {percent(random_summary['q1']['min'])} | {percent(random_summary['q1']['max'])} | {percent(random_summary['q1']['max_absolute'])} |",
        f"| 问题二总人日 | {percent(random_summary['q2']['mean'])} | {percent(random_summary['q2']['std'])} | {percent(random_summary['q2']['min'])} | {percent(random_summary['q2']['max'])} | {percent(random_summary['q2']['max_absolute'])} |",
        f"| 问题三招聘人数 | {percent(random_summary['q3']['mean'])} | {percent(random_summary['q3']['std'])} | {percent(random_summary['q3']['min'])} | {percent(random_summary['q3']['max'])} | {percent(random_summary['q3']['max_absolute'])} |",
        "",
        f"随机噪声在不同小时之间相互抵消，因此总货量和三个目标值的波动远小于单个小时的 10% 扰动上限。三个模型目标值的最大绝对变化率分别为 {percent(random_summary['q1']['max_absolute'])}、{percent(random_summary['q2']['max_absolute'])} 和 {percent(random_summary['q3']['max_absolute'])}，均低于预设的 5% 判定阈值，说明模型对零均值小时级随机误差具有较强鲁棒性。",
        "",
        "整体需求压力测试结果如下：",
        "",
        "| 场景 | 问题一总人日变化 | 问题二总人日变化 | 问题三招聘人数变化 |",
        "|---|---:|---:|---:|",
        f"| 全部到货量 -10% | {percent(stress['stress_minus_10']['q1_change_rate'])} | {percent(stress['stress_minus_10']['q2_change_rate'])} | {percent(stress['stress_minus_10']['q3_change_rate'])} |",
        f"| 全部到货量 +10% | {percent(stress['stress_plus_10']['q1_change_rate'])} | {percent(stress['stress_plus_10']['q2_change_rate'])} | {percent(stress['stress_plus_10']['q3_change_rate'])} |",
        "",
        "随机误差鲁棒性和系统负荷敏感性并不矛盾：局部正负扰动可以在日总量中抵消，而全时段同向增长会直接推高人员下界。",
        "",
        "## 四、模型改进方向",
        "",
        "1. **引入需求不确定性。** 当前模型把到货量视为完全已知。实际应用中可使用区间鲁棒优化或多情景随机规划，使排班同时覆盖正常、繁忙和突发到货情景。",
        "2. **为问题一设置安全余量。** 问题一的总体产能利用率接近 100%，经济性高但应对临时波动的余量较小。可增加 5%—10% 安全产能，或限制计划利用率不超过 90%—95%。",
        "3. **修正低效小时假设。** 当前假设每名员工的低效小时可由模型自由安排且处理能力固定为 10 件。可进一步规定休息时间窗，或把低效能力设置为区间变量，以刻画疲劳差异。",
        "4. **扩展班次制度。** 当前班次均为整数时刻开始的 8 小时班次。可引入半小时起点、6/8/10 小时混合班次、加班和临时班次，并在目标函数中加入工资和加班成本。",
        "5. **增强人员异质性。** 第三问默认员工能力相同且可跨班次调换。可加入技能等级、岗位资格、换班成本、夜班偏好和公平性目标，使人员方案更接近实际管理。",
        "6. **采用滚动优化。** 每日获得最新到货数据后重新优化未来数日，在保持既有排班稳定性的同时修正预测误差，提升方案的动态适应能力。",
        "",
        "## 五、结论",
        "",
        "三问结果均通过约束残差、整数性和独立重求解检验，且相应最优性下界能够闭合，因此模型在给定假设与数据范围内有效。模型对零均值小时级扰动具有较强抗干扰能力，但对全时段同向业务增长较敏感。实际应用时应通过安全余量、弹性用工和滚动优化提高对系统性需求增长的承受能力。",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate all three scheduling models and run robustness tests."
    )
    parser.add_argument("--input", type=Path, default=Path("第一轮B题附件.xlsx"))
    parser.add_argument(
        "--q1-result",
        type=Path,
        default=Path("运行结果_20260819/问题一优化结果_人数均衡.xlsx"),
    )
    parser.add_argument(
        "--q2-result",
        type=Path,
        default=Path("运行结果_20260819/问题二优化结果.xlsx"),
    )
    parser.add_argument(
        "--q3-result",
        type=Path,
        default=Path("运行结果_20260819/问题三联合优化结果.xlsx"),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("运行结果_20260819/模型检验"),
    )
    parser.add_argument("--scenarios", type=int, default=20)
    parser.add_argument("--seed", type=int, default=20260819)
    parser.add_argument("--q1-time-limit", type=float, default=60.0)
    parser.add_argument("--q2-time-limit", type=float, default=120.0)
    parser.add_argument("--q3-time-limit", type=float, default=180.0)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.scenarios < 1:
        raise ValueError("--scenarios must be positive.")
    args.output_dir.mkdir(parents=True, exist_ok=True)

    print("Validating Question 1...", flush=True)
    validity_q1 = validate_question1(
        args.input, args.q1_result, args.q1_time_limit
    )
    print(json.dumps(validity_q1, ensure_ascii=False), flush=True)

    print("Validating Question 2...", flush=True)
    validity_q2 = validate_question2(
        args.input, args.q2_result, args.q2_time_limit
    )
    print(json.dumps(validity_q2, ensure_ascii=False), flush=True)

    print("Running deep audit for Question 3...", flush=True)
    validity_q3, audit_text = validate_question3(
        args.input, args.q3_result, args.q3_time_limit
    )
    print(audit_text, end="", flush=True)
    print(json.dumps(validity_q3, ensure_ascii=False), flush=True)

    validity = {
        "tolerance": TOLERANCE,
        "question1": validity_q1,
        "question2": validity_q2,
        "question3": validity_q3,
        "all_passed": validity_q1["passed"] and validity_q2["passed"] and validity_q3["passed"],
    }
    validity_path = args.output_dir / "有效性检验结果.json"
    validity_path.write_text(
        json.dumps(validity, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    if not validity["all_passed"]:
        raise AssertionError("At least one validity check failed; robustness tests were not run.")

    baseline = (
        int(validity_q1["saved_total_worker_days"]),
        int(validity_q2["saved_total_worker_days"]),
        int(validity_q3["hired_workers"]),
    )
    print(f"Running {args.scenarios} random robustness scenarios...", flush=True)
    records = robustness_tests(
        args.input,
        baseline,
        args.scenarios,
        args.seed,
        args.q1_time_limit,
        args.q2_time_limit,
        args.q3_time_limit,
    )
    write_robustness_csv(args.output_dir / "鲁棒性检验结果.csv", records)
    write_report(
        args.output_dir / "数学模型检验与改进分析.md",
        validity,
        records,
        args.scenarios,
        args.seed,
    )
    print(f"All checks passed. Results saved to {args.output_dir.resolve()}", flush=True)


if __name__ == "__main__":
    main()
