"""Solve Question 3 under the processing conditions of Question 2.

Rules:
1. Every hired worker works exactly 23 days in the 30-day month.
2. No worker may work more than 7 consecutive days.
3. Every day must satisfy Question 2's deadline, low-rate-hour, and processing
   constraints; Question 2's optimized shift plan is not fixed.
4. Every day has five re-optimized shifts, and shifts may overlap.
5. A worker may work at most one shift per day and may change shifts by day.

The hiring model uses an aggregate state-flow formulation. A state records a
worker's current consecutive-work count (0..7) and total worked days (0..23).
Integer transition flows produce an exact minimum headcount without creating
hundreds of thousands of symmetric worker-level binary variables. The flows
are then decomposed into named worker schedules. Finally, each day's shifts,
10-item hours, and processing plan are re-optimized for the actual attendance.
"""

from __future__ import annotations

import argparse
import math
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.optimize import Bounds, LinearConstraint, milp
from scipy.sparse import coo_matrix

from solve_question2 import (
    EARLY_DEADLINE,
    EARLY_END,
    SHIFT_CAPACITY,
    STARTS,
    load_arrivals as load_hourly_arrivals,
    solve_day as solve_processing_day,
)


WORK_DAYS_PER_WORKER = 23
MAX_CONSECUTIVE = 7
SHIFT_LENGTH = 8
HOURS = 24
CONSECUTIVE_STATES = MAX_CONSECUTIVE + 1
WORK_TOTAL_STATES = WORK_DAYS_PER_WORKER + 1


@dataclass(frozen=True)
class DemandData:
    days: list[int]
    daily_minimum: np.ndarray
    shifts: dict[int, list[tuple[int, int, int]]]


@dataclass(frozen=True)
class FlowIndex:
    population: np.ndarray
    work: np.ndarray
    rest: np.ndarray
    hired: int
    max_surplus: int
    size: int


@dataclass
class HiringSolution:
    hired: int
    max_surplus: int
    daily_work_counts: np.ndarray
    population: np.ndarray
    work_flows: np.ndarray
    rest_flows: np.ndarray


def load_question2_demands(path: Path) -> DemandData:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if len(workbook.worksheets) < 2:
        raise ValueError("Question 2 workbook must contain summary and shift sheets.")

    summary = workbook.worksheets[0]
    daily_minimum: dict[int, int] = {}
    for row in summary.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        day = int(row[0])
        daily_minimum[day] = int(row[4])

    shift_sheet = workbook.worksheets[1]
    shifts: dict[int, list[tuple[int, int, int]]] = defaultdict(list)
    for row in shift_sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        day = int(row[0])
        start = int(row[2])
        end = int(row[3])
        workers = int(row[4])
        shifts[day].append((start, end, workers))

    days = sorted(daily_minimum)
    if not days:
        raise ValueError("No daily staffing demand was found.")
    if days != list(range(days[0], days[0] + len(days))):
        raise ValueError("Days in the Question 2 result must be consecutive.")
    for day in days:
        if len(shifts[day]) != 5:
            raise ValueError(f"Day {day} has {len(shifts[day])} shifts; expected 5.")
        if sum(workers for _, _, workers in shifts[day]) != daily_minimum[day]:
            raise ValueError(f"Day {day} shift demand does not match the daily summary.")
        shifts[day].sort()

    return DemandData(
        days=days,
        daily_minimum=np.array([daily_minimum[day] for day in days], dtype=int),
        shifts=dict(shifts),
    )


def make_flow_indices(day_count: int) -> FlowIndex:
    offset = 0
    population = np.arange(
        offset,
        offset + (day_count + 1) * CONSECUTIVE_STATES * WORK_TOTAL_STATES,
    ).reshape(day_count + 1, CONSECUTIVE_STATES, WORK_TOTAL_STATES)
    offset += (day_count + 1) * CONSECUTIVE_STATES * WORK_TOTAL_STATES
    work = np.arange(
        offset,
        offset + day_count * CONSECUTIVE_STATES * WORK_TOTAL_STATES,
    ).reshape(day_count, CONSECUTIVE_STATES, WORK_TOTAL_STATES)
    offset += day_count * CONSECUTIVE_STATES * WORK_TOTAL_STATES
    rest = np.arange(
        offset,
        offset + day_count * CONSECUTIVE_STATES * WORK_TOTAL_STATES,
    ).reshape(day_count, CONSECUTIVE_STATES, WORK_TOTAL_STATES)
    offset += day_count * CONSECUTIVE_STATES * WORK_TOTAL_STATES
    hired = offset
    offset += 1
    max_surplus = offset
    offset += 1
    return FlowIndex(
        population=population,
        work=work,
        rest=rest,
        hired=hired,
        max_surplus=max_surplus,
        size=offset,
    )


class SparseConstraints:
    def __init__(self, variable_count: int):
        self.variable_count = variable_count
        self.row_indices: list[int] = []
        self.column_indices: list[int] = []
        self.values: list[float] = []
        self.lower: list[float] = []
        self.upper: list[float] = []

    def add(self, coefficients: dict[int, float], lower: float, upper: float) -> None:
        row = len(self.lower)
        for column, value in coefficients.items():
            if value:
                self.row_indices.append(row)
                self.column_indices.append(column)
                self.values.append(float(value))
        self.lower.append(float(lower))
        self.upper.append(float(upper))

    def linear_constraint(self) -> LinearConstraint:
        matrix = coo_matrix(
            (self.values, (self.row_indices, self.column_indices)),
            shape=(len(self.lower), self.variable_count),
        ).tocsr()
        return LinearConstraint(matrix, np.array(self.lower), np.array(self.upper))


def maximum_work_days(length: int) -> int:
    return length - length // (MAX_CONSECUTIVE + 1)


def lower_bounds(daily_minimum: np.ndarray) -> dict[str, int]:
    day_count = len(daily_minimum)
    daily_bound = int(daily_minimum.max())
    total_bound = math.ceil(int(daily_minimum.sum()) / WORK_DAYS_PER_WORKER)
    window_bound = 0
    for start in range(day_count - MAX_CONSECUTIVE):
        demand = int(daily_minimum[start : start + MAX_CONSECUTIVE + 1].sum())
        window_bound = max(window_bound, math.ceil(demand / MAX_CONSECUTIVE))
    return {
        "max_daily": daily_bound,
        "total_workdays": total_bound,
        "eight_day_window": window_bound,
        "overall": max(daily_bound, total_bound, window_bound),
    }


def build_hiring_model(daily_minimum: np.ndarray) -> tuple[FlowIndex, SparseConstraints, Bounds, np.ndarray]:
    day_count = len(daily_minimum)
    idx = make_flow_indices(day_count)
    constraints = SparseConstraints(idx.size)

    bounds_info = lower_bounds(daily_minimum)
    upper_hired = max(bounds_info["overall"] * 2, bounds_info["overall"] + 100)

    variable_lower = np.zeros(idx.size, dtype=float)
    variable_upper = np.full(idx.size, float(upper_hired), dtype=float)
    variable_lower[idx.hired] = float(bounds_info["overall"])
    variable_upper[idx.hired] = float(upper_hired)
    variable_upper[idx.max_surplus] = float(upper_hired)

    # Invalid work transitions are fixed to zero.
    for day in range(day_count):
        for consecutive in range(CONSECUTIVE_STATES):
            for total_work in range(WORK_TOTAL_STATES):
                if consecutive >= MAX_CONSECUTIVE or total_work >= WORK_DAYS_PER_WORKER:
                    variable_upper[idx.work[day, consecutive, total_work]] = 0.0

    # Initial state: all hired workers have zero worked days and zero streak.
    for consecutive in range(CONSECUTIVE_STATES):
        for total_work in range(WORK_TOTAL_STATES):
            coefficients = {int(idx.population[0, consecutive, total_work]): 1.0}
            if consecutive == 0 and total_work == 0:
                coefficients[idx.hired] = -1.0
            constraints.add(coefficients, 0.0, 0.0)

    for day in range(day_count):
        # Every worker in a state either works or rests that day.
        for consecutive in range(CONSECUTIVE_STATES):
            for total_work in range(WORK_TOTAL_STATES):
                coefficients = {
                    int(idx.rest[day, consecutive, total_work]): 1.0,
                    int(idx.population[day, consecutive, total_work]): -1.0,
                }
                if consecutive < MAX_CONSECUTIVE and total_work < WORK_DAYS_PER_WORKER:
                    coefficients[int(idx.work[day, consecutive, total_work])] = 1.0
                constraints.add(coefficients, 0.0, 0.0)

        # Rest transitions reset the consecutive-work count to zero and keep
        # the accumulated monthly work count unchanged.
        for total_work in range(WORK_TOTAL_STATES):
            coefficients = {int(idx.population[day + 1, 0, total_work]): 1.0}
            for consecutive in range(CONSECUTIVE_STATES):
                coefficients[int(idx.rest[day, consecutive, total_work])] = -1.0
            constraints.add(coefficients, 0.0, 0.0)

        # Work transitions increase both state coordinates by one.
        for consecutive in range(1, CONSECUTIVE_STATES):
            for total_work in range(WORK_TOTAL_STATES):
                coefficients = {int(idx.population[day + 1, consecutive, total_work]): 1.0}
                if total_work >= 1:
                    coefficients[int(idx.work[day, consecutive - 1, total_work - 1])] = -1.0
                constraints.add(coefficients, 0.0, 0.0)

        daily_work_coefficients: dict[int, float] = {}
        for consecutive in range(MAX_CONSECUTIVE):
            for total_work in range(WORK_DAYS_PER_WORKER):
                daily_work_coefficients[int(idx.work[day, consecutive, total_work])] = 1.0
        constraints.add(daily_work_coefficients, float(daily_minimum[day]), np.inf)

        surplus_coefficients = dict(daily_work_coefficients)
        surplus_coefficients[idx.max_surplus] = -1.0
        constraints.add(surplus_coefficients, -np.inf, float(daily_minimum[day]))

    # Every hired worker must finish the month with exactly 23 worked days.
    final_coefficients = {idx.hired: -1.0}
    for consecutive in range(CONSECUTIVE_STATES):
        final_coefficients[int(idx.population[day_count, consecutive, WORK_DAYS_PER_WORKER])] = 1.0
    constraints.add(final_coefficients, 0.0, 0.0)

    integrality = np.ones(idx.size, dtype=int)
    return idx, constraints, Bounds(variable_lower, variable_upper), integrality


def solve_milp(
    objective: np.ndarray,
    constraints: SparseConstraints,
    bounds: Bounds,
    integrality: np.ndarray,
    time_limit: float,
    stage: str,
):
    result = milp(
        c=objective,
        integrality=integrality,
        bounds=bounds,
        constraints=constraints.linear_constraint(),
        options={
            "disp": False,
            "presolve": True,
            "time_limit": time_limit,
            "mip_rel_gap": 0.0,
        },
    )
    if result.status != 0 or result.x is None:
        raise RuntimeError(f"{stage} did not reach an optimum: status={result.status}, {result.message}")
    return result


def solve_hiring(daily_minimum: np.ndarray, time_limit: float) -> HiringSolution:
    idx, constraints, bounds, integrality = build_hiring_model(daily_minimum)

    objective = np.zeros(idx.size)
    objective[idx.hired] = 1.0
    stage1 = solve_milp(objective, constraints, bounds, integrality, time_limit, "Hiring stage")
    optimum_hired = int(round(float(stage1.x[idx.hired])))
    constraints.add({idx.hired: 1.0}, optimum_hired, optimum_hired)

    # Extra worker-days are unavoidable because every hired worker must work
    # exactly 23 days. Spread them by minimizing the largest daily surplus.
    objective = np.zeros(idx.size)
    objective[idx.max_surplus] = 1.0
    stage2 = solve_milp(objective, constraints, bounds, integrality, time_limit, "Surplus stage")
    optimum_surplus = int(round(float(stage2.x[idx.max_surplus])))
    constraints.add({idx.max_surplus: 1.0}, optimum_surplus, optimum_surplus)

    # Prefer solutions with fewer worker-days at a seven-day streak. This is a
    # tie-breaker only and cannot change hiring or daily surplus optima.
    objective = np.zeros(idx.size)
    for day in range(1, len(daily_minimum) + 1):
        objective[idx.population[day, MAX_CONSECUTIVE, :]] = 1.0
    stage3 = solve_milp(objective, constraints, bounds, integrality, time_limit, "Streak stage")

    values = stage3.x
    population = np.rint(values[idx.population]).astype(int)
    work_flows = np.rint(values[idx.work]).astype(int)
    rest_flows = np.rint(values[idx.rest]).astype(int)
    daily_work_counts = work_flows.sum(axis=(1, 2))

    solution = HiringSolution(
        hired=optimum_hired,
        max_surplus=optimum_surplus,
        daily_work_counts=daily_work_counts,
        population=population,
        work_flows=work_flows,
        rest_flows=rest_flows,
    )
    validate_aggregate_solution(solution, daily_minimum)
    return solution


def validate_aggregate_solution(solution: HiringSolution, daily_minimum: np.ndarray) -> None:
    if np.any(solution.daily_work_counts < daily_minimum):
        raise AssertionError("A day is staffed below its Question 2 requirement.")
    if int((solution.daily_work_counts - daily_minimum).max()) != solution.max_surplus:
        raise AssertionError("Maximum daily surplus is inconsistent.")
    if int(solution.daily_work_counts.sum()) != solution.hired * WORK_DAYS_PER_WORKER:
        raise AssertionError("Total worker-days do not equal 23 times the hired workers.")
    if int(solution.population[-1, :, WORK_DAYS_PER_WORKER].sum()) != solution.hired:
        raise AssertionError("Not every worker reaches exactly 23 worked days.")


def decompose_worker_days(solution: HiringSolution) -> tuple[list[list[int]], list[list[bool]]]:
    worker_count = solution.hired
    day_count = len(solution.daily_work_counts)
    state_members: dict[tuple[int, int], list[int]] = defaultdict(list)
    state_members[(0, 0)] = list(range(1, worker_count + 1))
    daily_workers: list[list[int]] = []
    work_calendar = [[False] * day_count for _ in range(worker_count + 1)]

    for day in range(day_count):
        next_states: dict[tuple[int, int], list[int]] = defaultdict(list)
        working_today: list[int] = []

        for consecutive in range(CONSECUTIVE_STATES):
            for total_work in range(WORK_TOTAL_STATES):
                members = sorted(state_members.get((consecutive, total_work), []))
                work_count = int(solution.work_flows[day, consecutive, total_work])
                rest_count = int(solution.rest_flows[day, consecutive, total_work])
                if work_count + rest_count != len(members):
                    raise AssertionError(f"Flow decomposition failed on day {day + 1}, state {(consecutive, total_work)}.")

                # Rotate the deterministic ordering by day to avoid assigning
                # the same IDs to every equivalent transition.
                if members:
                    rotation = day % len(members)
                    members = members[rotation:] + members[:rotation]
                workers = members[:work_count]
                resters = members[work_count:]

                if workers:
                    next_states[(consecutive + 1, total_work + 1)].extend(workers)
                    working_today.extend(workers)
                    for worker in workers:
                        work_calendar[worker][day] = True
                if resters:
                    next_states[(0, total_work)].extend(resters)

        if len(working_today) != int(solution.daily_work_counts[day]):
            raise AssertionError(f"Day {day + 1} worker decomposition count mismatch.")
        daily_workers.append(sorted(working_today))
        state_members = next_states

    for worker in range(1, worker_count + 1):
        if sum(work_calendar[worker]) != WORK_DAYS_PER_WORKER:
            raise AssertionError(f"Worker {worker} does not work exactly 23 days.")
        if maximum_consecutive(work_calendar[worker]) > MAX_CONSECUTIVE:
            raise AssertionError(f"Worker {worker} exceeds the consecutive-work limit.")
    return daily_workers, work_calendar


def allocate_daily_shift_counts(
    required_shifts: list[tuple[int, int, int]],
    scheduled_workers: int,
    time_limit: float,
) -> list[tuple[int, int, int, int]]:
    starts = [start for start, _, _ in required_shifts]
    ends = [end for _, end, _ in required_shifts]
    required = np.array([workers for _, _, workers in required_shifts], dtype=int)
    shift_count = len(required_shifts)

    # Variables: assigned shift counts, peak, and 24 scaled deviations.
    assigned_idx = np.arange(shift_count)
    peak_idx = shift_count
    deviation_idx = np.arange(shift_count + 1, shift_count + 1 + HOURS)
    variable_count = shift_count + 1 + HOURS

    rows: list[np.ndarray] = []
    lower: list[float] = []
    upper: list[float] = []

    row = np.zeros(variable_count)
    row[assigned_idx] = 1.0
    rows.append(row)
    lower.append(float(scheduled_workers))
    upper.append(float(scheduled_workers))

    coverage = np.array(
        [[int(start <= hour < start + SHIFT_LENGTH) for start in starts] for hour in range(HOURS)],
        dtype=float,
    )
    for hour in range(HOURS):
        row = np.zeros(variable_count)
        row[assigned_idx] = coverage[hour]
        row[peak_idx] = -1.0
        rows.append(row)
        lower.append(-np.inf)
        upper.append(0.0)

    variable_lower = np.zeros(variable_count)
    variable_lower[assigned_idx] = required
    variable_upper = np.full(variable_count, np.inf)
    integrality = np.ones(variable_count, dtype=int)
    bounds = Bounds(variable_lower, variable_upper)

    def run(objective: np.ndarray, extra_rows=None):
        use_rows = rows if extra_rows is None else rows + extra_rows[0]
        use_lower = lower if extra_rows is None else lower + extra_rows[1]
        use_upper = upper if extra_rows is None else upper + extra_rows[2]
        result = milp(
            c=objective,
            integrality=integrality,
            bounds=bounds,
            constraints=LinearConstraint(np.vstack(use_rows), np.array(use_lower), np.array(use_upper)),
            options={"disp": False, "time_limit": time_limit, "mip_rel_gap": 0.0},
        )
        if result.status != 0 or result.x is None:
            raise RuntimeError(f"Daily extra-shift allocation failed: {result.message}")
        return result

    objective = np.zeros(variable_count)
    objective[peak_idx] = 1.0
    stage1 = run(objective)
    optimum_peak = int(round(float(stage1.x[peak_idx])))

    extra_rows: list[np.ndarray] = []
    extra_lower: list[float] = []
    extra_upper: list[float] = []
    row = np.zeros(variable_count)
    row[peak_idx] = 1.0
    extra_rows.append(row)
    extra_lower.append(float(optimum_peak))
    extra_upper.append(float(optimum_peak))

    for hour in range(HOURS):
        row_above = np.zeros(variable_count)
        row_above[deviation_idx[hour]] = 1.0
        row_above[assigned_idx] = -3.0 * coverage[hour]
        extra_rows.append(row_above)
        extra_lower.append(-float(scheduled_workers))
        extra_upper.append(np.inf)

        row_below = np.zeros(variable_count)
        row_below[deviation_idx[hour]] = 1.0
        row_below[assigned_idx] = 3.0 * coverage[hour]
        extra_rows.append(row_below)
        extra_lower.append(float(scheduled_workers))
        extra_upper.append(np.inf)

    objective = np.zeros(variable_count)
    objective[deviation_idx] = 1.0
    stage2 = run(objective, (extra_rows, extra_lower, extra_upper))
    assigned = np.rint(stage2.x[assigned_idx]).astype(int)
    return [(starts[i], ends[i], int(required[i]), int(assigned[i])) for i in range(shift_count)]


def assign_workers_to_shifts(
    daily_workers: list[list[int]],
    demand: DemandData,
    time_limit: float,
) -> tuple[dict[int, list[str]], dict[int, list[tuple[int, int, int, int]]]]:
    worker_count = max(worker for workers in daily_workers for worker in workers)
    calendar = {worker: ["休"] * len(demand.days) for worker in range(1, worker_count + 1)}
    worker_shift_counts: dict[int, dict[int, int]] = {
        worker: defaultdict(int) for worker in range(1, worker_count + 1)
    }
    assigned_by_day: dict[int, list[tuple[int, int, int, int]]] = {}

    for day_index, day in enumerate(demand.days):
        allocations = allocate_daily_shift_counts(
            demand.shifts[day],
            len(daily_workers[day_index]),
            time_limit,
        )
        assigned_by_day[day] = allocations
        remaining = set(daily_workers[day_index])

        # Assign larger shifts first, while balancing each worker's history of
        # the same start time.
        for start, end, _, assigned_count in sorted(allocations, key=lambda item: -item[3]):
            candidates = sorted(
                remaining,
                key=lambda worker: (worker_shift_counts[worker][start], worker),
            )
            chosen = candidates[:assigned_count]
            if len(chosen) != assigned_count:
                raise AssertionError(f"Day {day}: not enough workers for shift {start}-{end}.")
            for worker in chosen:
                calendar[worker][day_index] = f"{start:02d}-{end:02d}"
                worker_shift_counts[worker][start] += 1
                remaining.remove(worker)

        if remaining:
            raise AssertionError(f"Day {day}: {len(remaining)} working workers were not assigned a shift.")
    return calendar, assigned_by_day


def maximum_consecutive(work_flags: list[bool]) -> int:
    maximum = 0
    current = 0
    for worked in work_flags:
        if worked:
            current += 1
            maximum = max(maximum, current)
        else:
            current = 0
    return maximum


def style_worksheet(worksheet, widths: dict[int, float]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = width


def write_results(
    path: Path,
    demand: DemandData,
    solution: HiringSolution,
    bounds_info: dict[str, int],
    calendar: dict[int, list[str]],
    assigned_by_day: dict[int, list[tuple[int, int, int, int]]],
) -> None:
    workbook = Workbook()

    summary = workbook.active
    summary.title = "招聘汇总"
    summary.append(["指标", "数值", "说明"])
    summary_rows = [
        ("最少招聘人数", solution.hired, "第三问目标值"),
        ("每人工作天数", WORK_DAYS_PER_WORKER, "每名工人恰好工作23天"),
        ("总排班人日", solution.hired * WORK_DAYS_PER_WORKER, "招聘人数×23"),
        ("第二问最低需求人日", int(demand.daily_minimum.sum()), "30天最低需求合计"),
        ("额外排班人日", solution.hired * WORK_DAYS_PER_WORKER - int(demand.daily_minimum.sum()), "由每人工作23天产生"),
        ("最大单日额外人数", solution.max_surplus, "在最少招聘人数下进一步最小化"),
        ("最大日需求下界", bounds_info["max_daily"], "任何一天都必须满足"),
        ("总人日下界", bounds_info["total_workdays"], "最低需求人日÷23"),
        ("连续8天下界", bounds_info["eight_day_window"], "任意8天每人最多工作7天"),
        ("综合理论下界", bounds_info["overall"], "以上下界最大值"),
    ]
    for row in summary_rows:
        summary.append(row)

    daily = workbook.create_sheet("每日人员")
    daily.append(["天", "第二问最低人数", "第三问安排人数", "额外人数", "休息人数", "班次安排"])
    for index, day in enumerate(demand.days):
        allocations = assigned_by_day[day]
        labels = [f"{start:02d}-{end:02d}:{assigned}" for start, end, _, assigned in allocations]
        scheduled = int(solution.daily_work_counts[index])
        minimum = int(demand.daily_minimum[index])
        daily.append([day, minimum, scheduled, scheduled - minimum, solution.hired - scheduled, "；".join(labels)])

    coverage = workbook.create_sheet("班次覆盖")
    coverage.append(["天", "开始时间", "结束时间", "第二问需求", "第三问安排", "额外人数", "是否满足"])
    for day in demand.days:
        for start, end, required, assigned in assigned_by_day[day]:
            coverage.append([day, start, end, required, assigned, assigned - required, "是" if assigned >= required else "否"])

    worker_stats = workbook.create_sheet("人员检验")
    worker_stats.append(["工人编号", "工作天数", "休息天数", "最大连续工作天数", "是否合格"])
    for worker, entries in calendar.items():
        flags = [entry != "休" for entry in entries]
        worked = sum(flags)
        max_run = maximum_consecutive(flags)
        valid = worked == WORK_DAYS_PER_WORKER and max_run <= MAX_CONSECUTIVE
        worker_stats.append([worker, worked, len(entries) - worked, max_run, "是" if valid else "否"])

    monthly = workbook.create_sheet("人员月历")
    monthly.append(["工人编号"] + [f"第{day}天" for day in demand.days] + ["工作天数", "最大连续"])
    for worker, entries in calendar.items():
        flags = [entry != "休" for entry in entries]
        monthly.append([worker] + entries + [sum(flags), maximum_consecutive(flags)])

    details = workbook.create_sheet("逐人上班明细")
    details.append(["工人编号", "天", "班次"])
    for worker, entries in calendar.items():
        for day_index, entry in enumerate(entries):
            if entry != "休":
                details.append([worker, demand.days[day_index], entry])

    style_worksheet(summary, {1: 22, 2: 15, 3: 34})
    style_worksheet(daily, {1: 8, 2: 17, 3: 17, 4: 12, 5: 12, 6: 65})
    style_worksheet(coverage, {1: 8, 2: 12, 3: 12, 4: 15, 5: 15, 6: 12, 7: 12})
    style_worksheet(worker_stats, {1: 12, 2: 12, 3: 12, 4: 18, 5: 12})
    style_worksheet(monthly, {1: 12, **{column: 11 for column in range(2, 32)}, 32: 12, 33: 12})
    style_worksheet(details, {1: 12, 2: 8, 3: 12})
    monthly.freeze_panes = "B2"

    workbook.save(path)


def derive_daily_minimums(
    arrivals_by_day: dict[int, np.ndarray],
    time_limit: float,
) -> tuple[DemandData, dict[int, object]]:
    """Project Question 2's conditions to a minimum headcount for each day.

    These minima are used only as necessary daily lower bounds in the monthly
    hiring model. Their shift starts and shift-level staffing are not fixed in
    Question 3.
    """
    minimum_solutions: dict[int, object] = {}
    shifts: dict[int, list[tuple[int, int, int]]] = {}
    days = sorted(arrivals_by_day)
    daily_minimum: list[int] = []

    for day in days:
        solution = solve_processing_day(day, arrivals_by_day[day], time_limit)
        minimum_solutions[day] = solution
        daily_minimum.append(solution.total_workers)
        day_shifts: list[tuple[int, int, int]] = []
        for position in np.flatnonzero(solution.selected):
            start = STARTS[position]
            day_shifts.append((start, start + SHIFT_LENGTH, int(solution.workers_by_start[position])))
        shifts[day] = sorted(day_shifts)
        print(f"minimum day={day:02d} workers={solution.total_workers}")

    return (
        DemandData(days=days, daily_minimum=np.array(daily_minimum, dtype=int), shifts=shifts),
        minimum_solutions,
    )


def solve_question3_daily_schedules(
    arrivals_by_day: dict[int, np.ndarray],
    days: list[int],
    daily_work_counts: np.ndarray,
    time_limit: float,
) -> dict[int, object]:
    """Re-optimize each day under Question 2 conditions and Q3 attendance."""
    solutions: dict[int, object] = {}
    for day_index, day in enumerate(days):
        fixed_workers = int(daily_work_counts[day_index])
        solution = solve_processing_day(
            day,
            arrivals_by_day[day],
            time_limit,
            fixed_workers=fixed_workers,
        )
        if solution.total_workers != fixed_workers:
            raise AssertionError(f"Day {day}: fixed-worker processing solution changed the attendance count.")
        solutions[day] = solution
        starts = [STARTS[position] for position in np.flatnonzero(solution.selected)]
        print(f"question3 day={day:02d} workers={fixed_workers} starts={starts}")
    return solutions


def assign_workers_to_joint_shifts(
    daily_workers: list[list[int]],
    days: list[int],
    daily_solutions: dict[int, object],
) -> dict[int, list[str]]:
    worker_count = max(worker for workers in daily_workers for worker in workers)
    calendar = {worker: ["休"] * len(days) for worker in range(1, worker_count + 1)}
    worker_shift_counts: dict[int, dict[int, int]] = {
        worker: defaultdict(int) for worker in range(1, worker_count + 1)
    }

    for day_index, day in enumerate(days):
        solution = daily_solutions[day]
        remaining = set(daily_workers[day_index])
        shift_counts = []
        for position in np.flatnonzero(solution.selected):
            start = STARTS[position]
            shift_counts.append(
                (start, start + SHIFT_LENGTH, int(solution.workers_by_start[position]))
            )

        if sum(count for _, _, count in shift_counts) != len(remaining):
            raise AssertionError(f"Day {day}: daily processing shifts do not match working personnel.")

        for start, end, count in sorted(shift_counts, key=lambda item: -item[2]):
            candidates = sorted(
                remaining,
                key=lambda worker: (worker_shift_counts[worker][start], worker),
            )
            chosen = candidates[:count]
            if len(chosen) != count:
                raise AssertionError(f"Day {day}: not enough workers for shift {start}-{end}.")
            for worker in chosen:
                calendar[worker][day_index] = f"{start:02d}-{end:02d}"
                worker_shift_counts[worker][start] += 1
                remaining.remove(worker)

        if remaining:
            raise AssertionError(f"Day {day}: unassigned working personnel remain.")
    return calendar


def write_joint_results(
    path: Path,
    arrivals_by_day: dict[int, np.ndarray],
    demand: DemandData,
    hiring: HiringSolution,
    bounds_info: dict[str, int],
    daily_solutions: dict[int, object],
    calendar: dict[int, list[str]],
) -> None:
    workbook = Workbook()

    summary = workbook.active
    summary.title = "招聘汇总"
    summary.append(["指标", "数值", "说明"])
    rows = [
        ("最少招聘人数", hiring.hired, "第三问目标值"),
        ("每人工作天数", WORK_DAYS_PER_WORKER, "每名工人恰好工作23天"),
        ("总排班人日", hiring.hired * WORK_DAYS_PER_WORKER, "招聘人数×23"),
        ("满足问题二条件的最低需求人日", int(demand.daily_minimum.sum()), "仅作为必要下界，不固定第二问班次结论"),
        ("第三问额外排班人日", hiring.hired * WORK_DAYS_PER_WORKER - int(demand.daily_minimum.sum()), "由每人工作23天产生"),
        ("最大单日额外人数", hiring.max_surplus, "在最少招聘人数下进一步最小化"),
        ("最大日需求下界", bounds_info["max_daily"], "由问题二条件逐日求得"),
        ("总人日下界", bounds_info["total_workdays"], "最低需求人日÷23"),
        ("连续8天下界", bounds_info["eight_day_window"], "任意8天每人最多工作7天"),
        ("综合理论下界", bounds_info["overall"], "以上下界最大值"),
    ]
    for row in rows:
        summary.append(row)

    daily = workbook.create_sheet("每日人员与检验")
    daily.append(
        [
            "天",
            "问题二条件下最低人数",
            "第三问实际人数",
            "额外人数",
            "休息人数",
            "16点早期剩余",
            "日末总剩余",
            "第三问重新优化的班次",
        ]
    )
    for day_index, day in enumerate(demand.days):
        solution = daily_solutions[day]
        labels = []
        for position in np.flatnonzero(solution.selected):
            start = STARTS[position]
            labels.append(f"{start:02d}-{start + SHIFT_LENGTH:02d}:{int(solution.workers_by_start[position])}")
        actual = int(hiring.daily_work_counts[day_index])
        minimum = int(demand.daily_minimum[day_index])
        daily.append(
            [
                day,
                minimum,
                actual,
                actual - minimum,
                hiring.hired - actual,
                round(float(solution.inventory_early[EARLY_DEADLINE - 1]), 6),
                round(float(solution.inventory_early[-1] + solution.inventory_late[-1]), 6),
                "；".join(labels),
            ]
        )

    shifts = workbook.create_sheet("第三问班次")
    shifts.append(["天", "班次序号", "开始时间", "结束时间", "工人数", "班次理论能力"])
    low_hours = workbook.create_sheet("10件小时安排")
    low_hours.append(["天", "班次开始", "班次内小时", "全局小时", "按10件工作的工人数"])
    for day in demand.days:
        solution = daily_solutions[day]
        for order, position in enumerate(np.flatnonzero(solution.selected), start=1):
            start = STARTS[position]
            count = int(solution.workers_by_start[position])
            shifts.append([day, order, start, start + SHIFT_LENGTH, count, SHIFT_CAPACITY * count])
            for local_hour in range(SHIFT_LENGTH):
                low_hours.append(
                    [
                        day,
                        start,
                        local_hour + 1,
                        start + local_hour,
                        int(solution.low_hour_workers[position, local_hour]),
                    ]
                )

    hourly = workbook.create_sheet("逐小时处理检验")
    hourly.append(
        [
            "天",
            "小时",
            "进货量",
            "早期处理",
            "普通处理",
            "总处理",
            "早期剩余",
            "普通剩余",
            "总剩余",
            "在岗人数",
            "其中10件人数",
            "最大处理能力",
            "未使用产能",
        ]
    )
    for day in demand.days:
        solution = daily_solutions[day]
        for hour in range(HOURS):
            processed = float(solution.processed_early[hour] + solution.processed_late[hour])
            inventory = float(solution.inventory_early[hour] + solution.inventory_late[hour])
            hourly.append(
                [
                    day,
                    hour,
                    round(float(arrivals_by_day[day][hour]), 6),
                    round(float(solution.processed_early[hour]), 6),
                    round(float(solution.processed_late[hour]), 6),
                    round(processed, 6),
                    round(float(solution.inventory_early[hour]), 6),
                    round(float(solution.inventory_late[hour]), 6),
                    round(inventory, 6),
                    int(solution.hourly_workers[hour]),
                    int(solution.hourly_low_workers[hour]),
                    int(solution.hourly_capacity[hour]),
                    round(float(solution.hourly_capacity[hour] - processed), 6),
                ]
            )

    worker_stats = workbook.create_sheet("人员检验")
    worker_stats.append(["工人编号", "工作天数", "休息天数", "最大连续工作天数", "是否合格"])
    for worker, entries in calendar.items():
        flags = [entry != "休" for entry in entries]
        worked = sum(flags)
        max_run = maximum_consecutive(flags)
        valid = worked == WORK_DAYS_PER_WORKER and max_run <= MAX_CONSECUTIVE
        worker_stats.append([worker, worked, len(entries) - worked, max_run, "是" if valid else "否"])

    monthly = workbook.create_sheet("人员月历")
    monthly.append(["工人编号"] + [f"第{day}天" for day in demand.days] + ["工作天数", "最大连续"])
    for worker, entries in calendar.items():
        flags = [entry != "休" for entry in entries]
        monthly.append([worker] + entries + [sum(flags), maximum_consecutive(flags)])

    details = workbook.create_sheet("逐人上班明细")
    details.append(["工人编号", "天", "班次"])
    for worker, entries in calendar.items():
        for day_index, entry in enumerate(entries):
            if entry != "休":
                details.append([worker, demand.days[day_index], entry])

    style_worksheet(summary, {1: 28, 2: 15, 3: 42})
    style_worksheet(daily, {1: 8, 2: 22, 3: 18, 4: 12, 5: 12, 6: 16, 7: 14, 8: 70})
    style_worksheet(shifts, {1: 8, 2: 12, 3: 12, 4: 12, 5: 12, 6: 16})
    style_worksheet(low_hours, {1: 8, 2: 12, 3: 14, 4: 12, 5: 22})
    style_worksheet(hourly, {1: 8, 2: 9, 3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 12, 9: 12, 10: 12, 11: 14, 12: 16, 13: 14})
    style_worksheet(worker_stats, {1: 12, 2: 12, 3: 12, 4: 18, 5: 12})
    style_worksheet(monthly, {1: 12, **{column: 11 for column in range(2, 32)}, 32: 12, 33: 12})
    style_worksheet(details, {1: 12, 2: 8, 3: 12})
    monthly.freeze_panes = "B2"
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Solve Question 3 under Question 2 processing conditions.")
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("第一轮B题附件.xlsx"),
        help="Original hourly arrival workbook",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("问题三联合优化结果.xlsx"),
        help="Question 3 output workbook",
    )
    parser.add_argument(
        "--time-limit",
        type=float,
        default=180.0,
        help="MILP time limit in seconds for each stage",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    arrivals_by_day = load_hourly_arrivals(args.input)
    demand, _ = derive_daily_minimums(arrivals_by_day, args.time_limit)
    bounds_info = lower_bounds(demand.daily_minimum)
    print(f"Lower bounds: {bounds_info}")

    solution = solve_hiring(demand.daily_minimum, args.time_limit)
    print(
        f"Optimal hired workers={solution.hired}, "
        f"maximum daily surplus={solution.max_surplus}, "
        f"total worker-days={int(solution.daily_work_counts.sum())}"
    )

    daily_workers, _ = decompose_worker_days(solution)
    daily_solutions = solve_question3_daily_schedules(
        arrivals_by_day,
        demand.days,
        solution.daily_work_counts,
        args.time_limit,
    )
    calendar = assign_workers_to_joint_shifts(daily_workers, demand.days, daily_solutions)
    write_joint_results(
        args.output,
        arrivals_by_day,
        demand,
        solution,
        bounds_info,
        daily_solutions,
        calendar,
    )
    print(f"Saved result to: {args.output.resolve()}")


if __name__ == "__main__":
    main()
